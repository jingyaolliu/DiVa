# 用于生成valid score(看哪些词的valid score最高) -- 再考虑要不要在此基础上加上bleu score
# 将12位放novelty-score
import numpy as np
import openpyxl
import random
import json
import xlsxwriter
from sklearn.decomposition import PCA
from openpyxl.styles import PatternFill
import torch
from scipy import stats
# pca=PCA(n_components=1)
import torch.nn.functional as F

device=torch.device('cuda' if torch.cuda.is_available() else 'cpu')
green_fill = PatternFill(fgColor="AACF91", fill_type="solid")
red_fill=PatternFill(fgColor="F9966B", fill_type="solid")
blue_fill=PatternFill(fgColor="9AFEFF", fill_type="solid")
yellow_fill=PatternFill(fgColor="FFFFCC", fill_type="solid")
# 灰色
grey_fill=PatternFill(fgColor="D3D3D3", fill_type="solid")


# annotated_dict[song_name]['annotated_labels_test']=annotated_labels
# annotated_dict[song_name]['precision_test']=p
# annotated_dict[song_name]['recall_test']=r
# annotated_dict[song_name]['f1_test']=f1
# annotated_dict[song_name]['loss_annotation_num_test']=loss_annotation_num

# bold font
from openpyxl.styles import Font
bold_font = Font(bold=True,color='cc3300')
normal_font=Font(bold=False)

# 加上bleu
# from soft_match_metrics_bleu import *

all_vec_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/tme_big_data/candidate_words_context_embedding_dict_1536_300_pca512.npy'
vec_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/tme_big_data/golden_dict_1536_300_pca512.npy'
# golden_dict_xlnet.npy
# vec_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/tme_big_data/golden_dict_xlnet.npy'
vec_dict=np.load(vec_path,allow_pickle=True).item()
# 所有golden labels 
# todo 用来查看扩增的golden labels
golden_labels=list(vec_dict.keys())

# all label 频率字典
label_freq=json.load(open('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Evaluate/yao_check_datas/words_tf_info_freq.json'))
# all_candiartes=list(label_freq.keys())
#todo valid score - 1词频的方差 2离散系数
def cal_valid_score(label_freq):
    # label : valid score
    max_val=0
    min_val=1
    for key in label_freq.keys():
        tf_infos=np.array(label_freq[key]['freq'])
        # valid_score=np.var(tf_infos)
        valid_score=np.std(tf_infos)/np.mean(tf_infos)
        label_freq[key]['var']=valid_score
        if valid_score>max_val:
            max_val=valid_score
        if valid_score<min_val:
            min_val=valid_score
    # max-min normalization
    for key in label_freq.keys():
        label_freq[key]['var']=(label_freq[key]['var']-min_val)/(max_val-min_val)

    return label_freq
# label_freq=cal_valid_score(label_freq)
# wirte label freq to excel
def write_label_freq_to_excel(label_freq):
    workbook=openpyxl.Workbook()
    sheet=workbook.active
    sheet.title='label_freq'
    sheet.cell(1,1).value='label'
    sheet.cell(1,2).value='var/valid score'
    # sort 
    label_freq=sorted(label_freq.items(),key=lambda x:x[1]['var'],reverse=True)
    for i in range(len(label_freq)):
        sheet.cell(i+2,1).value=label_freq[i][0]
        sheet.cell(i+2,2).value=label_freq[i][1]['var']
        # sheet.cell(i+2,3).value=label_freq[i][1]['dn']
    workbook.save('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Evaluate/yao_check_datas/0323_label_freq(valid).xlsx')
# write_label_freq_to_excel(label_freq)

# keys=['ferq','dn']
def cal_all_valid_label(label_freq):
    valid_num=0
    # label : valid score
    label_dict={}
    for key in label_freq.keys():
        max_freq=max(label_freq[key]['freq'])
        mean_freq=np.mean(label_freq[key]['freq'])
        valid_score=max_freq-mean_freq
        label_dict[key]=valid_score
        if valid_score>0.0068:
            valid_num+=1
    # sort by valid score
    label_dict=sorted(label_dict.items(),key=lambda x:x[1],reverse=True)
    print('valid_percentage:',valid_num/len(label_freq.keys()))

# cal_all_valid_label(label_freq)
# 
all_vec_dict=np.load(all_vec_path,allow_pickle=True).item()
save_folder='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Evaluate/yao_check_datas/iter_check/'

# 测试文件
# test_f_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Evaluate/eval_data/annotation_ebc_val.json'
# todo 文件名中有(focal)的是用focal loss矫正过的数据 - ebc(focal)_val
train_f_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/eval_data/train_ebc_val.json'
test_f_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/eval_data/ebc_val.json'
origin_f_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/ebc_data/finetune_test_data_1123.json'

# 原始文件
origin_train_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/train_t0_for_supervise_DIVA_1123.json'
origin_test_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/test_t0_for_supervise_DIVA_1123.json'
object_train_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/train_t0_for_supervise_DIVA_0312.json'
object_test_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/test_t0_for_supervise_DIVA_0312.json'


# centers文件地址
# centers_vecs_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Evaluate/yao_check_datas/k-means/centers_vecs.npy'
# centers_vecs=np.load(centers_vecs_path,allow_pickle=True)

song_info_dict=json.load(open('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Evaluate/yao_check_datas/song_info_dict.json','r'))

# cal cosine similarity
def cos_sim(vec1, vec2):
    vec1=np.array(vec1)
    vec2=np.array(vec2)
    return np.dot(vec1, vec2) / (np.linalg.norm(vec1) * np.linalg.norm(vec2))

def cos_sim_torch(mat1 ,mat2):
    """
    mat1 vecs: candidate_num * dimension
    mat2 centers: N * dimension
    
    to calculate the cosine similarity between two matrixs
    """
    # cos_sim=F.cosine_similarity(mat1,mat2,dim=1)
    
    mat1=mat1/torch.norm(mat1,dim=1,keepdim=True)
    mat2=mat2/torch.norm(mat2,dim=1,keepdim=True)
    cos_sim=mat1 @ mat2.T
    return cos_sim

# 计算与centers最短的距离（10词求mean）
def mean_center_dist(vecs,centers):
    # centers = times(10),k(15),dimension(512)
    """
    vecs: candidate_num * dimension
    centers: M * N * dimension
    """
    # list2torch
    all_labels_dist=[]
    vecs=vecs.to(device)
    # 对vecs进行切割
    vecs_split=torch.split(vecs,20000,dim=0)
    
    new_centers=[]
    for centers_m in centers:
        new_centers_tmp=[]
        for center in centers_m:
            center=torch.tensor(center)
            new_centers_tmp.append(center)
        new_centers.append(torch.stack(new_centers_tmp).to(device))
        
    # centers=torch.stack(new_centers).to(device)
    # [candidate_num, dim] to [1,candidate_num,dim] unsqueeze
    for vecs in vecs_split:
        # vecs=vecs.unsqueeze(0)
        # transpose
        # centers=centers.transpose(1,2)
        # 计算cosine similarity
        # cos_sim_torch=torch.nn.CosineSimilarity(dim=2,eps=1e-6)
        # 计算与centers的距离
        all_dist_tmp=[]
        for centers in new_centers:
            dist=cos_sim_torch(vecs,centers)
            # 计算最大值
            dist=torch.max(dist,dim=1)[0]
            # squeeze
            # dist=dist.squeeze(0)
            # 计算每一行的平均值
            all_dist_tmp.append(dist)
        
        all_dist_tmp_mean=torch.mean(torch.stack(all_dist_tmp),dim=0)
        all_dist_tmp_mean=all_dist_tmp_mean.cpu().tolist()
        # dist=cos_sim_torch(vecs,centers)
        # # 计算最大值
        # dist=torch.max(dist,dim=-1)[0]
        # # 计算每一行的平均值
        # dist=torch.mean(dist,dim=0)
        # # torch to list(cpu)
        # dist=dist.cpu().tolist()
        all_labels_dist+=all_dist_tmp_mean
    # for i in range(len(centers)):
    #     # 当前center集合
    #     max_dist=0
    #     tmp_centers=centers[i]
    #     for center in tmp_centers:
    #         dist=cos_sim(vec,center)
    #         if dist>max_dist:
    #             max_dist=dist
    #     dist_lst.append(max_dist)
    # dist_lst=np.array(dist_lst)
    print(np.array(all_labels_dist).mean())
    return all_labels_dist
    
def oder_valid_in_golden(vec_dict,f_name):
    workbook=openpyxl.Workbook()
    worksheet=workbook.active
    line=1

    for key in vec_dict.keys():
        row=1
        worksheet.cell(line,row,key)
        row+=1
        # {label:cos_sim}
        sim_dict={}
        for key2 in vec_dict.keys():
            if key!=key2:
                sim_dict[key2]=cos_sim(vec_dict[key],vec_dict[key2])
        # sort
        sim_dict=sorted(sim_dict.items(),key=lambda x:x[1],reverse=True)
        for i in range(30):
            worksheet.cell(line,row,sim_dict[i][0])
            worksheet.cell(line+1,row,sim_dict[i][1])
            row+=1
        line+=2
    
    # save 
    workbook.save(save_folder+f_name+'.xlsx')

# oder_valid_in_golden(vec_dict,'valid_oder_top30(xlnet)')

# 获取分数排名字典
def get_score_dict(score_dict,site):
    score_dict=sorted(score_dict.items(),key=lambda x:x[1][site],reverse=True)
    score_dict2={item[0]:i+1 for i,item in enumerate(score_dict)}
    return score_dict2

# 生成valid score(test 集 candidate中的valid score) output top5 similar golden label
def valid_score(all_vec_dict,vec_dict,test_f_path,object_excel,golden_labels=golden_labels,label_freq=label_freq,song_info_dict=song_info_dict,N=16372):
    # print(len(song_info_dict))
    # added_golden_num 计算扩增的golden labels
    broaden_golden_labels_top10=[]
    broaden_golden_labels_top15=[]
    broaden_labels=[]
    song_lables=[]
    # song_name 与 predict_labels
    test_song_infos={}
    # 判断数据pearson相关系数
    pearson_rs=[]

    #统计每首歌新增的labels

    workbook=openpyxl.Workbook()
    worksheet=workbook.active
    worksheet.title='annotation'
    line=1
    data_lst=json.load(open(test_f_path,'r'))
    # data_lst_origin=json.load(open(origin_f_path,'r'))
    new_data_lst=[]

    f1_lst=[]
    
    # analysis data index list
    # index_lst=[]
    #todo 获取一个词的所有ebc 分数 
    # label:[ebc_s_lst]
    ebc_s_dict={}
    # novelty_s_lst=[]
    # tf_idf_s_lst=[]
    # diverse_s_lst=[]
    for i in range(1000):
        data_item = data_lst[i]
        score_dict=data_item['final_score_dict_sort']
        for label in score_dict.keys():
            if label not in ebc_s_dict:
                ebc_s_dict[label]=[]
            ebc_s_dict[label].append(score_dict[label][-1])
    #         diverse_s_lst.append(score_dict[label][4]*score_dict[label][12])
    #         novelty_s_lst.append(score_dict[label][12])
    #         tf_idf_s_lst.append(score_dict[label][4])
    # diverse_s_max=max(diverse_s_lst)
    # diverse_s_min=min(diverse_s_lst)
    # novelty_s_max=max(novelty_s_lst)
    # novelty_s_min=min(novelty_s_lst)
    # tf_idf_s_max=max(tf_idf_s_lst)
    # tf_idf_s_min=min(tf_idf_s_lst)


    for j in range(50):
        if j%20==0:
            print('------------------{}----------------------'.format(j))
        # while True:
        #     index=random.randint(0,len(data_lst)-1)
        #     if index not in index_lst:
        #         index_lst.append(index)
        #         data_item=data_lst[index]
        #         break
        index=j
        data_item=data_lst[index]
        # data_item_origin=data_lst_origin[index]
        # data_item['song_comments_detail_final']=data_item_origin['song_comments_detail_final']
        song_name=data_item['song_name']
        #todo annotated labels
        # if song_name not in song_info_dict:
        #     print('{}:{}'.format(index,song_name))
        #     continue
        # annotated_labels=song_info_dict[song_name]['annotated_labels']
        song_golden_labels=data_item['song_pseudo_golden_labels']
        if 'song_all_silver_labels' in data_item:
            song_silver_labels=data_item['song_all_silver_labels']
        else:
            song_silver_labels=[]
        song_silver_labels_no_golden=list(set(song_silver_labels)-set(song_golden_labels))
        # song_lables=list(set(data_item['song_pseudo_golden_labels']+data_item['annotation_labels']))
        score_dict=data_item['final_score_dict_sort']
        song_all_labels=list(set(song_golden_labels+song_silver_labels))

        # get ebc rank dict
        ebc_rank_dict=get_score_dict(score_dict,-1)
        freq_rank_dict=get_score_dict(score_dict,4)
        novelty_rank_dict=get_score_dict(score_dict,12)

        worksheet.cell(line,1,song_name)
        line+=1 

        #comment num
        worksheet.cell(line,1,'comment num')
        worksheet.cell(line,2,len(data_item['song_comments_detail_final']))
        line+=1

        # write golden labels
        worksheet.cell(line,1,'golden labels').fill=red_fill
        worksheet.cell(line,2,' '.join(list(data_item['song_pseudo_golden_labels'])))
        line+=1
        worksheet.cell(line,1,'song labels')
        worksheet.cell(line,2,' '.join(song_all_labels))
        line+=1

        worksheet.cell(line,1,'candidate num')
        worksheet.cell(line,2,len(score_dict))
        worksheet.cell(line,3,0.05*len(score_dict))
        line+=1

        pca_origin_lst=[]

        # candidate labels | golden labels(valid score)
        # for key in score_dict.keys():
        #     row=1
        #     worksheet.cell(line,row,key)
        #     row+=1

            #? # 10-15分类计算 valid score
            # candidate_vec=all_vec_dict[key]
            # valid_score=mean_center_dist(candidate_vec)
            # score_dict[key][7]=cal_bleu(key,golden_labels,2)
            # pca_origin_lst.append(np.array([score_dict[key][5],score_dict[key][7]]))

            # 将新的valid score 写入site 6
            # score_dict[key][6]=valid_score
            # 将bleu score 写入site 
            # old valid score+new valid score -> site 8
            # score_dict[key][7]=(score_dict[key][5]+score_dict[key][6])/2
        # sort by old valid score in site 5
        # new_pca_datas=pca.fit_transform(np.array(pca_origin_lst))
        # wite pca data in site 8
        # index=0
        # for key in score_dict.keys():
        #     score_dict[key][8]=new_pca_datas[index][0]
        #     index+=1
            
        # sort_score_dict1=sorted(score_dict.items(),key=lambda x:x[1][6],reverse=True)
        # # max-min normalize
        # max_score=sort_score_dict1[0][1][6]
        # min_score=sort_score_dict1[-1][1][6]
        ebc_vals=[]
        novelty_vals=[]
        half_candidate_num=len(score_dict)/2
        ebc_more_than_095=0
        for key in score_dict.keys():
            if score_dict[key][-1]>=0.95:
                ebc_more_than_095+=1
            # max-min normalize
            # tmp_s=(score_dict[key][6]-min_score)/(max_score-min_score)
            # tmp_s=score_dict[key][6]
            # 提取两边的词
            # score_dict[key][6]=math.pow(0.5-tmp_s,2)
            # max-min normalize
            # score_dict[key][4]=(score_dict[key][4]-tf_idf_s_min)/(tf_idf_s_max-tf_idf_s_min)
            # score_dict[key][12]=(score_dict[key][12]-novelty_s_min)/(novelty_s_max-novelty_s_min)
            #todo site 6 : ebc_punsish(根据loss来判断easy or hard)
            #todo site 6 :  entropy_uncertainty
            label_uncertainty=-(score_dict[key][-1]*math.log(score_dict[key][-1])+(1-score_dict[key][-1])*math.log((1-score_dict[key][-1])))
            score_dict[key][6]=label_uncertainty
            # if key in song_all_labels:
            #     score_dict[key][6]=1-score_dict[key][-1]
            # else:
            #     score_dict[key][6]=score_dict[key][-1]
            #todo site 9: tf-idf++ with ebc_rank
            score_dict[key][9]=score_dict[key][4]/math.sqrt(ebc_rank_dict[key])
            #todo site 10: 1-valid score(new)+ebc_score
            score_dict[key][10]=math.pow(score_dict[key][12],2)+math.pow(score_dict[key][-1],2)
            #todo valid score tf_max-tf_mean

            #-标签扩充思路更新------novelty 搭配 tf-idf++来扩充标签
            # score_dict[key][13]=score_dict[key][4]/math.sqrt(novelty_rank_dict[key])
            # score_dict[key][13]=score_dict[key][12]/math.sqrt(freq_rank_dict[key])
            # score_dict[key][13]=score_dict[key][12]/math.log(freq_rank_dict[key]+1)
            score_dict[key][13]=score_dict[key][12]*score_dict[key][4]
            # tf_info=np.array(label_freq[key]['freq'])
            # score_dict[key][11]=tf_info.max()-tf_info.mean()
            # score_dict[key][11]=score_dict[key][10]/math.log(freq_rank_dict[key]+1,2)
            # score_dict[key][12]=score_dict[key][10]/math.sqrt(freq_rank_dict[key])
            ebc_vals.append(score_dict[key][-1])
            novelty_vals.append(score_dict[key][12])
        
        # sort by site 6
        site_6_rank_dict=get_score_dict(score_dict,6)
        # todo 根据valid score 收紧标注规则
        #获取 site 10的rank
        # ebc_novelty_rank_dict=get_score_dict(score_dict,10)
        #todo tf-idf++/ebc_novelty_rank -- fusion_score
        #todo 判断ebc是否呈正态分布
        # ebc_vals=np.array(ebc_vals)
        # # log变化 偏态转正态
        # ebc_vals=np.log(ebc_vals+1)
        # u=ebc_vals.mean()
        # sigma=ebc_vals.var()
        # # result=stats.kstest(ebc_vals,'norm',(u,sigma))
        # # 偏度
        # R_sc=stats.skew(ebc_vals)
        # R_ku=stats.kurtosis(ebc_vals)
        # # 峰度
        # print('偏度：{}   峰度：{}'.format(R_sc,R_ku))
        

        # for key in score_dict.keys():
        #     score_dict[key][11]=score_dict[key][4]/math.sqrt(ebc_novelty_rank_dict[key])
        # # sort by site 11
        # sort_score_dict_fusion=sorted(score_dict.items(),key=lambda x:x[1][11],reverse=True)
        # max_fu_score=sort_score_dict_fusion[0][1][11]
        # min_fu_score=sort_score_dict_fusion[-1][1][11]
        # for i in range(len(sort_score_dict_fusion)):
        #     sort_score_dict_fusion[i][1][11]=(sort_score_dict_fusion[i][1][11]-min_fu_score)/(max_fu_score-min_fu_score)

        # # sort by site 12
        # sort_score_dict_fusion2=sorted(score_dict.items(),key=lambda x:x[1][12],reverse=True)
        # max_fu_score2=sort_score_dict_fusion2[0][1][12]
        # min_fu_score2=sort_score_dict_fusion2[-1][1][12]
        # for i in range(len(sort_score_dict_fusion2)):
        #     sort_score_dict_fusion2[i][1][12]=(sort_score_dict_fusion2[i][1][12]-min_fu_score2)/(max_fu_score2-min_fu_score2)
        
        # new_annotated_labels=[]
        # for label in annotated_labels:
        #     if score_dict[label][11]>0.003:
        #         new_annotated_labels.append(label)
        #todo ebc_vals 与 novelty_vals之间的pearson相关系数
        # pearson_r=np.corrcoef(np.array(ebc_vals),np.array(novelty_vals))[0][1]
        # pearson_rs.append(pearson_r)
        #todo sort by novelty score 
        # sort_score_dict1=sorted(score_dict.items(),key=lambda x:x[1][6],reverse=True)
        
        # sort by ebc
        sort_score_dict_ebc=sorted(score_dict.items(),key=lambda x:x[1][-1],reverse=True)

        #todo sort by novelty_tf-idf++
        sort_score_dict_novelty_tf=sorted(score_dict.items(),key=lambda x:x[1][13],reverse=True)
        # max_novelty_tf_score=sort_score_dict_novelty_tf[0][1][13]
        # min_novelty_tf_score=sort_score_dict_novelty_tf[-1][1][13]
        # max-min normalize
        for i in range(len(sort_score_dict_novelty_tf)):
            # score_dict[sort_score_dict_novelty_tf[i][0]][13]=(sort_score_dict_novelty_tf[i][1][13]-diverse_s_min)/(diverse_s_max-diverse_s_min)
            # sort_score_dict_novelty_tf[i][1][13]=(sort_score_dict_novelty_tf[i][1][13]-diverse_s_min)/(diverse_s_max-diverse_s_min)
            #todo new_fusion ebc_val+new_valid+diverse_置信度
            ebc_infos=np.array(ebc_s_dict[sort_score_dict_novelty_tf[i][0]])
            length_punish=min(1,len(sort_score_dict_novelty_tf[i][0])/2)
            new_valid=(label_freq[sort_score_dict_novelty_tf[i][0]]['var']+ebc_infos.mean())*length_punish*0.5
            score_dict[sort_score_dict_novelty_tf[i][0]][5]=new_valid
        # sort by site 14 new_fusion
        new_valid_rank_dict=get_score_dict(score_dict,5)
        # sort_score_dict_new_fusion=sorted(score_dict.items(),key=lambda x:x[1][5],reverse=True)
        #todo get new fusion
        for key in score_dict.keys():
            score_dict[key][8]=score_dict[key][13]/math.log(new_valid_rank_dict[key]+1,2)
        sort_score_dict_new_fusion=sorted(score_dict.items(),key=lambda x:x[1][8],reverse=True)
        # max_fu_score=sort_score_dict_new_fusion[0][1][8]
        # min_fu_score=sort_score_dict_new_fusion[-1][1][8]
        # # max-min normalize
        # for i in range(len(sort_score_dict_new_fusion)):
        #     sort_score_dict_new_fusion[i][1][8]=(sort_score_dict_new_fusion[i][1][8]-min_fu_score)/(max_fu_score-min_fu_score)
        # sort by site 8 new_fusion
        # sort_score_dict_new_fusion=sorted(score_dict.items(),key=lambda x:x[1][8],reverse=True)
        
        

        # sort by ebc score valid score fusion in site 9
        sort_score_dict_freq=sorted(score_dict.items(),key=lambda x:x[1][9],reverse=True)
        max_fu_score=sort_score_dict_freq[0][1][9]
        min_fu_score=sort_score_dict_freq[-1][1][9]

        # max-min normalize
        for i in range(len(sort_score_dict_freq)):
            sort_score_dict_freq[i][1][9]=(sort_score_dict_freq[i][1][9]-min_fu_score)/(max_fu_score-min_fu_score)

        # sort by valid score fusion in site 10
        sort_score_dict_valid=sorted(score_dict.items(),key=lambda x:x[1][10],reverse=True)


        # ebc分数大于0.95
        # label_ebc_more_than_095_lst=[]
        # for i in range(len(sort_score_dict_ebc)):
        #     if sort_score_dict_ebc[i][1][-1]>=0.95:
        #         label_ebc_more_than_095_lst.append(sort_score_dict_ebc[i][0])
        #     else:
        #         break

        # # novelty score 大于0.9
        # label_novelty_match_more_than_09_lst=[]
        # for i in range(len(sort_score_dict_valid)):
        #     if sort_score_dict_valid[i][1][10]>=0.9:
        #         if len(sort_score_dict_valid[i][0])>1 and sort_score_dict_valid[i][1][11]>0.0068:
        #             label_novelty_match_more_than_09_lst.append(sort_score_dict_valid[i][0])
        #     else:
        #         break
        
        # # mactch score with tf-idf++ 前5%取整
        # candidates_num=int(len(score_dict)*0.05)
        # label_match_tf_idf_top5_lst=[]
        # for i in range(candidates_num):
        #     if len(sort_score_dict_freq[i][0])>1 and sort_score_dict_freq[i][1][11]>0.0068:
        #         label_match_tf_idf_top5_lst.append(sort_score_dict_freq[i][0])


        # pred_labels=list(set(label_ebc_more_than_095_lst+label_novelty_match_more_than_09_lst+label_match_tf_idf_top5_lst))
        # test_song_infos[song_name]={}
        # test_song_infos[song_name]['pred_labels']=pred_labels
        # # test_song_infos[song_name]['annotated_labels']=new_annotated_labels
        # test_song_infos[song_name]['golden_labels']=song_golden_labels

        # # true_labels=list(set(new_annotated_labels+song_golden_labels))
        # # p=len(list(set(pred_labels)&set(true_labels)))/len(pred_labels)
        # # r=len(list(set(pred_labels)&set(true_labels)))/len(true_labels)
        # # f1=2*p*r/(p+r)
        # # f1_lst.append(f1)
        # broaden_labels+=pred_labels
        # todo 计算f1
        


        # # 若ebc分数大于0.95的标签数小于15，则补充其他标签
        # # 最少的valid+ebc修正标签数 15
        # len_label_valid_amend=15
        # if len(label_ebc_more_than_095_lst)<15:
        #     len_label_valid_amend+= (15-len(label_ebc_more_than_095_lst))

        # label_valid_amend_lst=[]
        # for i in range(len_label_valid_amend):
        #     label_valid_amend_lst.append(sort_score_dict_valid[i][0])
        # # 补充标签 0.9以上的
        # if len(label_ebc_more_than_095_lst)<15:
        #     tmp_site=len_label_valid_amend
        #     while(sort_score_dict_valid[tmp_site][1][10]>=0.9):
        #         label_valid_amend_lst.append(sort_score_dict_valid[tmp_site][0])
        #         tmp_site+=1
        # label_valid_amend_lst_annotation=[]
        # # valid sort 0.7以上均作为备选label
        # for i in range(len(sort_score_dict_valid)):
        #     if sort_score_dict_valid[i][1][10]>=0.7:
        #         label_valid_amend_lst_annotation.append(sort_score_dict_valid[i][0])
        #     else:
        #         break
        # # 0.5 以上- 扩大标注范围
        # label_valid_amend_lst_annotation_2=[]
        # for i in range(len(sort_score_dict_valid)):
        #     if sort_score_dict_valid[i][1][10]>=0.5:
        #         label_valid_amend_lst_annotation_2.append(sort_score_dict_valid[i][0])
        #     else:
        #         break
        
            

        # pred_labels_top15=[]
        # # consist of ebc,valid,freq
        # pred_labels_top15_ebc=[]
        # # todo 目前暂时用到这个
        # pred_labels_top15_freq=[]
        # pred_labels_top15_valid=[]
        # i=0
        # while len(pred_labels_top15_ebc)<15:
        #     if sort_score_dict_ebc[i][0] not in pred_labels_top15_ebc and len(sort_score_dict_ebc[i][0])>1:
        #         pred_labels_top15_ebc.append(sort_score_dict_ebc[i][0])
        #     i+=1
        # i=0
        # while len(pred_labels_top15_freq)<15:
        #     if sort_score_dict_freq[i][0] not in pred_labels_top15_freq:
        #         pred_labels_top15_freq.append(sort_score_dict_freq[i][0])
        #     i+=1
        # i=0
        # while len(pred_labels_top15_valid)<15:
        #     if sort_score_dict_valid[i][0] not in pred_labels_top15_valid and len(sort_score_dict_valid[i][0])>1:
        #         pred_labels_top15_valid.append(sort_score_dict_valid[i][0])
        #     i+=1
        # pred_labels_top15=list(set(pred_labels_top15_ebc+pred_labels_top15_freq+pred_labels_top15_valid))

        # #! add pred_top15 fusion sets
        # data_item['pred_labels_top15']=pred_labels_top15

        # #! 当前预测结果
        # pred_labels=list(set(label_valid_amend_lst+label_ebc_more_than_095_lst+pred_labels_top15_freq))
        # new_pred_labels=[]
        # # 删除其中长度为1 的标签
        # for i in range(len(pred_labels)):
        #     if len(pred_labels[i])>1:
        #         new_pred_labels.append(pred_labels[i])
        # data_item['model_pred_labels']=new_pred_labels
        # broaden_labels+=pred_labels
        
        # # broaden_golden_labels_top10+=pred_labels_top10
        # broaden_golden_labels_top15+=pred_labels_top15

        # pred_labels_top20=[]
        # # consist of ebc,valid,freq
        # pred_labels_top20_ebc=[]
        # pred_labels_top20_freq=[]
        # pred_labels_top20_valid=[]
        # i=0
        # while len(pred_labels_top20_ebc)<20:
        #     if sort_score_dict_ebc[i][0] not in pred_labels_top20_ebc and len(sort_score_dict_ebc[i][0])>1:
        #         pred_labels_top20_ebc.append(sort_score_dict_ebc[i][0])
        #     i+=1
        # i=0
        # while len(pred_labels_top20_freq)<20:
        #     if sort_score_dict_freq[i][0] not in pred_labels_top20_freq:
        #         pred_labels_top20_freq.append(sort_score_dict_freq[i][0])
        #     i+=1
        # i=0
        # while len(pred_labels_top20_valid)<20:
        #     if sort_score_dict_valid[i][0] not in pred_labels_top20_valid and len(sort_score_dict_valid[i][0])>1:
        #         pred_labels_top20_valid.append(sort_score_dict_valid[i][0])
        #     i+=1
        # pred_labels_top20=list(set(pred_labels_top20_ebc+pred_labels_top20_freq+pred_labels_top20_valid))
        
        

        # # for annotation
        # pred_labels_top30_valid=[]
        # # consist of ebc,valid,freq
        # pred_labels_top30_ebc=[]
        # pred_labels_top30_freq=[]
        # pred_labels_top30_valid=[]
        # i=0
        # while len(pred_labels_top30_valid)<30:
        #     if sort_score_dict_valid[i][0] not in pred_labels_top30_valid and len(sort_score_dict_valid[i][0])>1:
        #         pred_labels_top30_valid.append(sort_score_dict_valid[i][0])
        #     i+=1
        # i=0
        # while len(pred_labels_top30_ebc)<30:
        #     if sort_score_dict_ebc[i][0] not in pred_labels_top30_ebc:
        #         pred_labels_top30_ebc.append(sort_score_dict_ebc[i][0])
        #     i+=1
        # i=0
        # while len(pred_labels_top30_freq)<30:
        #     if sort_score_dict_freq[i][0] not in pred_labels_top30_freq and len(sort_score_dict_freq[i][0])>1:
        #         pred_labels_top30_freq.append(sort_score_dict_freq[i][0])
        #     i+=1
        # pred_labels_top30=list(set(pred_labels_top30_ebc+pred_labels_top30_freq+pred_labels_top30_valid))
        
        # #! 当前标注候选范围
        # annotation_candidate_labels=list(set(label_valid_amend_lst_annotation+label_ebc_more_than_095_lst+pred_labels_top30_freq))
        # new_annotation_candidate_labels=[]
        # # 删除其中长度为1 的标签
        # for i in range(len(annotation_candidate_labels)):
        #     if len(annotation_candidate_labels[i])>1:
        #         new_annotation_candidate_labels.append(annotation_candidate_labels[i])
        # data_item['annotation_candidate_labels']=new_annotation_candidate_labels
        
        # #! 当前更大的标注候选范围
        # annotation_candidate_labels_2=list(set(label_valid_amend_lst_annotation_2+label_ebc_more_than_095_lst+pred_labels_top30_freq))
        # new_annotation_candidate_labels_2=[]
        # for i in range(len(annotation_candidate_labels_2)):
        #     if len(annotation_candidate_labels_2[i])>1:
        #         new_annotation_candidate_labels_2.append(annotation_candidate_labels_2[i])
        # data_item['annotation_candidate_labels_2']=new_annotation_candidate_labels_2


        
        # new_data_lst.append(data_item)





        # line+=1
        # worksheet.cell(line,1,'candidate labels').fill=green_fill
        # worksheet.cell(line,2,'、'.join(list(set())))

        # # fusion predict label top 25
        # fu_pred_labels_top25=[]
        # while len(fu_pred_labels_top25)<25:
        #     if sort_score_dict_fusion[i][0] not in fu_pred_labels_top25 and len(sort_score_dict_fusion[i][0])>1:
        #         fu_pred_labels_top25.append(sort_score_dict_fusion[i][0])
        #     i+=1

        # worksheet.cell(line,1,'predict_label_fused top 25').fill=blue_fill
        # worksheet.cell(line,2,len(fu_pred_labels_top25))
        # worksheet.cell(line,3,'、'.join(fu_pred_labels_top25))
        # line+=1
        
        # 每首歌之间有间隔

        # line+=1


        # detail info

        
        #todo write in worksheet
        # # ebc
        row=1
        worksheet.cell(line,row,'candidate')
        row+=1
        worksheet.cell(line,row,'ebc score')
        row+=1
        worksheet.cell(line,row,'rf-idf++')
        row+=1
        worksheet.cell(line,row,'novelty_score')
        row+=1
        worksheet.cell(line,row,'valid_score')
        row+=2

        # rf-idf++
        # row=1
        # worksheet.cell(line,row,'candidate')
        # row+=1
        # worksheet.cell(line,row,'rf-idf++ score')
        # row+=1
        # worksheet.cell(line,row,'ebc_score')
        # row+=2

        #freq with ebc_rank
        worksheet.cell(line,row,'candidate')
        row+=1
        # 因为这里使用了ebc_rank，所以一定要最后进行max-min normalize
        worksheet.cell(line,row,'rf-idf++/sqrt(ebc_rank) score')
        row+=1
        worksheet.cell(line,row,'ebc_score')
        row+=1
        worksheet.cell(line,row,'novelty_score')
        row+=1
        worksheet.cell(line,row,'valid_score')
        row+=2

        #valid with ebc score 两者充当相互补充的位置
        worksheet.cell(line,row,'candidate')
        row+=1
        worksheet.cell(line,row,'novelty_score^2+ebc_score^2')
        row+=1
        worksheet.cell(line,row,'ebc_score')
        row+=1
        worksheet.cell(line,row,'tf-idf++_score')
        row+=1
        # worksheet.cell(line,row,'tf-idf++')
        # row+=1
        # worksheet.cell(line,row,'tf-idf')
        # row+=1
        # worksheet.cell(line,row,'dn')
        # row+=1
        # worksheet.cell(line,row,'tn')
        # row+=1
        # worksheet.cell(line,row,'tn_mean')
        # row+=1
        # worksheet.cell(line,row,'tn_max')
        # row+=1
        # worksheet.cell(line,row,'tn_min')
        # row+=1
        worksheet.cell(line,row,'valid_score')
        row+=2

        # worksheet.cell(line,row,'candidate')
        # row+=1
        # worksheet.cell(line,row,'fusion_score')
        # row+=1
        # worksheet.cell(line,row,'ebc_score')
        # row+=1
        # worksheet.cell(line,row,'valid_score')
        # row+=2

        # worksheet.cell(line,row,'candidate')
        # row+=1
        # worksheet.cell(line,row,'fusion_score2')
        # row+=1
        # worksheet.cell(line,row,'ebc_score')
        # row+=1
        # worksheet.cell(line,row,'valid_score')
        # row+=2
        # line+=1
        worksheet.cell(line,row,'candidate')
        row+=1
        worksheet.cell(line,row,'tf-idf++/sqrt(novely_rank) score')
        row+=1
        worksheet.cell(line,row,'ebc_score')
        row+=1
        worksheet.cell(line,row,'valid_score')
        row+=1
        worksheet.cell(line,row,'ebc_infos(max)')
        row+=1
        worksheet.cell(line,row,'ebc_infos(min)')
        row+=1
        worksheet.cell(line,row,'ebc_infos(mean)')
        row+=2

        worksheet.cell(line,row,'candidate')
        row+=1
        worksheet.cell(line,row,'new_fusion')
        row+=2



        # new_valid_score
        # worksheet.cell(line,12,'candidate')
        # worksheet.cell(line,13,'new valid score')
        # worksheet.cell(line,14,'tf-idf++')
        # worksheet.cell(line,15,'tf-idf')
        # worksheet.cell(line,16,'all counts')
        # worksheet.cell(line,17,'dn')
        line+=1

        # # # 直接将三个分数融合在一起
        # # worksheet.cell(line,10,'candidate')
        # # worksheet.cell(line,11,'rf-idf++ valid ebc score fusion')
        # # worksheet.cell(line,12,'ebc_valid/ sqrt(rf-idf++ rank)')
        # # line+=1

        # tmp_line = line
        # # write new valid score in excel
        # for i in range(len(score_dict)):
        #     if sort_score_dict1[i][0] in song_lables:
        #         worksheet.cell(line,1,sort_score_dict1[i][0]).fill=red_fill
        #     else:
        #         worksheet.cell(line,1,sort_score_dict1[i][0])
        #     worksheet.cell(line,2,sort_score_dict1[i][1][6])
        #     worksheet.cell(line,3,sort_score_dict1[i][1][4])
        #     worksheet.cell(line,4,sort_score_dict1[i][1][0])
        #     line+=1

        candidate_num=min(int(0.05*len(score_dict)),(0.05*len(score_dict)+ebc_more_than_095)/2)
        # 记录ebc分数大于0.95的标签数量
        ebc_more_than_095=0
        

        row=1
        tmp_line=line
        # write ebc score in excel
        for i in range(len(score_dict)):
            if score_dict[sort_score_dict_ebc[i][0]][-1]>=0.95:
                ebc_more_than_095+=1
            tmp_row=row
            if sort_score_dict_ebc[i][1][-1] >= 0.95:
                worksheet.cell(line,tmp_row,sort_score_dict_ebc[i][0]).fill=red_fill
            else:
                worksheet.cell(line,tmp_row,sort_score_dict_ebc[i][0])
            tmp_row+=1
            worksheet.cell(line,tmp_row,sort_score_dict_ebc[i][1][-1])
            tmp_row+=1
            worksheet.cell(line,tmp_row,sort_score_dict_ebc[i][1][4])
            tmp_row+=1
            worksheet.cell(line,tmp_row,sort_score_dict_ebc[i][1][12])
            tmp_row+=1
            worksheet.cell(line,tmp_row,label_freq[sort_score_dict_ebc[i][0]]['var'])
            tmp_row+=2
            line+=1
        row=tmp_row

        # write rf-idf++ score in excel'
        # line=tmp_line
        # score_dict_sort_4=sorted(score_dict.items(),key=lambda x:x[1][4],reverse=True)
        # for i in range(len(score_dict_sort_4)):
        #     tmp_row=row
        #     if score_dict_sort_4[i][1][-1] <= 0.01 and score_dict_sort_4[i][1][-1]>0.0001:
        #         worksheet.cell(line,tmp_row,score_dict_sort_4[i][0]).fill=red_fill
        #     else:
        #         worksheet.cell(line,tmp_row,score_dict_sort_4[i][0])
        #     tmp_row+=1
        #     worksheet.cell(line,tmp_row,score_dict_sort_4[i][1][4])
        #     tmp_row+=1
        #     worksheet.cell(line,tmp_row,score_dict_sort_4[i][1][-1])
        #     tmp_row+=2
        #     line+=1
        # row=tmp_row

        
        line=tmp_line
        # write freq score in excel
        for i in range(len(score_dict)):
            tmp_row=row
            # print(row)
            # valid info
            tf_infos=np.array(label_freq[sort_score_dict_freq[i][0]]['freq'])
            # dn=sort_score_dict_valid[i][1][2]
            tn_mean=tf_infos.mean()
            tn_max=tf_infos.max()
            dn=sort_score_dict_freq[i][1][2]
            # tn_min=tf_infos.min()
            #todo valid score 为方差
            valid_score=label_freq[sort_score_dict_freq[i][0]]['var']
            # valid_score=(tn_max-tn_mean)/min(0.5*math.log(dn+1)*math.log(N/(dn+1)),math.log(dn+1))
            # valid info
            # todo 查看多少标签是无关ebc能力被提出的
            if sort_score_dict_freq[i][1][-1] < 0.05 and sort_score_dict_freq[i][1][-1] > 0.0001 and i < (candidate_num+ebc_more_than_095)/2 and valid_score >= 0.14:
                worksheet.cell(line,tmp_row,sort_score_dict_freq[i][0]).fill=red_fill
            else:
                worksheet.cell(line,tmp_row,sort_score_dict_freq[i][0])
            # if i==candidate_num-1:
            #     worksheet.cell(line,tmp_row,sort_score_dict_freq[i][0]).fill=red_fill
            tmp_row+=1
            worksheet.cell(line,tmp_row,sort_score_dict_freq[i][1][9])
            tmp_row+=1
            # if valid_score < 0.001:
            #     worksheet.cell(line,tmp_row,valid_score).fill=grey_fill
            # else:
            #     worksheet.cell(line,tmp_row,valid_score)
            worksheet.cell(line,tmp_row,sort_score_dict_freq[i][1][-1])
            tmp_row+=1
            worksheet.cell(line,tmp_row,sort_score_dict_freq[i][1][12])
            tmp_row+=1
            worksheet.cell(line,tmp_row,valid_score)
            tmp_row+=2
            line+=1
        row=tmp_row

        # wirte in worksheet
        line=tmp_line
        for i in range(len(score_dict)):
            tmp_row=row
            # valid score
            tf_infos=np.array(label_freq[sort_score_dict_valid[i][0]]['freq'])
            dn=sort_score_dict_valid[i][1][2]
            tn_mean=tf_infos.mean()
            tn_max=tf_infos.max()
            tn_min=tf_infos.min()
            #todo valid score 为方差
            # valid_score=np.var(tf_infos)   
            valid_score=label_freq[sort_score_dict_valid[i][0]]['var'] 
            # valid_score=(tn_max-tn_mean)/min(0.5*math.log(dn+1)*math.log(N/(dn+1)),math.log(dn+1))
            # valid score
            #todo 查看多少标签是无关ebc能力被提出的
            if sort_score_dict_valid[i][1][-1] < 0.05 and sort_score_dict_valid[i][1][-1] > 0.0001 and sort_score_dict_valid[i][1][10] >=0.98 and valid_score >= 0.14:
                worksheet.cell(line,tmp_row,sort_score_dict_valid[i][0]).fill=red_fill
            else:
                worksheet.cell(line,tmp_row,sort_score_dict_valid[i][0])
            tmp_row+=1
            worksheet.cell(line,tmp_row,sort_score_dict_valid[i][1][10])
            tmp_row+=1
            worksheet.cell(line,tmp_row,sort_score_dict_valid[i][1][-1])
            tmp_row+=1
            worksheet.cell(line,tmp_row,sort_score_dict_valid[i][1][4])
            tmp_row+=1
            # worksheet.cell(line,tmp_row,sort_score_dict_valid[i][1][3])
            # tmp_row+=1
            # worksheet.cell(line,tmp_row,sort_score_dict_valid[i][1][2])
            # tmp_row+=1
            # worksheet.cell(line,tmp_row,sort_score_dict_valid[i][1][1])
            # tmp_row+=1
            # worksheet.cell(line,tmp_row,tf_infos.mean())
            # tmp_row+=1
            # worksheet.cell(line,tmp_row,tf_infos.max())
            # tmp_row+=1
            # worksheet.cell(line,tmp_row,tf_infos.min())
            # tmp_row+=1
            # if valid_score < 0.001:
            #     worksheet.cell(line,tmp_row,valid_score).fill=grey_fill
            # else:
            #     worksheet.cell(line,tmp_row,valid_score)
            worksheet.cell(line,tmp_row,valid_score)
            tmp_row+=2
            line+=1
        row=tmp_row

        line=tmp_line
        for i in range(len(score_dict)):
            ebc_infos=np.array(ebc_s_dict[sort_score_dict_novelty_tf[i][0]])
            lenth_punish=min(1,len(sort_score_dict_novelty_tf[i][0])/2)
            ebc_max=ebc_infos.max()
            ebc_min=ebc_infos.min()
            ebc_mean=ebc_infos.mean()
            old_valid=label_freq[sort_score_dict_novelty_tf[i][0]]['var']
            new_valid=lenth_punish*(ebc_mean+old_valid)
            tmp_row=row
            # novelty_tf-idf++
            worksheet.cell(line,tmp_row,sort_score_dict_novelty_tf[i][0])
            tmp_row+=1
            worksheet.cell(line,tmp_row,sort_score_dict_novelty_tf[i][1][13])
            tmp_row+=1
            worksheet.cell(line,tmp_row,sort_score_dict_novelty_tf[i][1][-1])
            tmp_row+=1
            worksheet.cell(line,tmp_row,label_freq[sort_score_dict_novelty_tf[i][0]]['var'])
            tmp_row+=1
            worksheet.cell(line,tmp_row,ebc_infos.max())
            tmp_row+=1
            worksheet.cell(line,tmp_row,ebc_infos.min())
            tmp_row+=1
            worksheet.cell(line,tmp_row,ebc_infos.mean())
            tmp_row+=1
            worksheet.cell(line,tmp_row,new_valid)
            tmp_row+=2
            line+=1
        
        row=tmp_row

        line=tmp_line
        for i in range(len(score_dict)):
            tmp_row=row
            tmp_font=normal_font
            # if sort_score_dict_new_fusion[i][0] in song_silver_labels_no_golden and site_6_rank_dict[sort_score_dict_new_fusion[i][0]]/len(score_dict)<0.2:
            #     tmp_font=bold_font
            if site_6_rank_dict[sort_score_dict_new_fusion[i][0]]/len(score_dict)<0.5 and i<candidate_num:
                worksheet.cell(line,tmp_row,sort_score_dict_new_fusion[i][0]).fill=red_fill
                # worksheet.cell(line,tmp_row,sort_score_dict_new_fusion[i][0]).font=tmp_font
            else:
                worksheet.cell(line,tmp_row,sort_score_dict_new_fusion[i][0])
            worksheet.cell(line,tmp_row,sort_score_dict_new_fusion[i][0])
            tmp_row+=1
            worksheet.cell(line,tmp_row,sort_score_dict_new_fusion[i][1][8])
            tmp_row+=1
            worksheet.cell(line,tmp_row,sort_score_dict_new_fusion[i][1][-1])
            tmp_row+=1
            worksheet.cell(line,tmp_row,sort_score_dict_new_fusion[i][1][6])
            tmp_row+=1
            worksheet.cell(line,tmp_row,site_6_rank_dict[sort_score_dict_new_fusion[i][0]]/len(score_dict))
            tmp_row+=2
            line+=1
        
        row=tmp_row
        # # write new valid score in excel
        # line=tmp_line
        # for i in range(len(score_dict)):
        #     if sort_score_dict1[i][0] in song_lables:
        #         worksheet.cell(line,12,sort_score_dict1[i][0]).fill=red_fill
        #     else:
        #         worksheet.cell(line,12,sort_score_dict1[i][0])
        #     worksheet.cell(line,13,sort_score_dict1[i][1][6])
        #     worksheet.cell(line,14,sort_score_dict1[i][1][4])
        #     # worksheet.cell(line,15,golden_label_freq[sort_score_dict1[i][0]])
        #     worksheet.cell(line,15,sort_score_dict1[i][1][2])
        #     line+=1
        
        # # wirte in worksheet fusion data
        # line=tmp_line
        # for i in range(len(score_dict)):
        #     if sort_score_dict_fusion[i][0] in song_lables:
        #         worksheet.cell(line,10,sort_score_dict_fusion[i][0]).fill=yellow_fill
        #     else:
        #         worksheet.cell(line,10,sort_score_dict_fusion[i][0])
        #     worksheet.cell(line,11,sort_score_dict_fusion[i][1][11])
        #     line+=1
        # # wirte in worksheet fusion data
        # diverse_candidate_num=len(score_dict)*0.05
        # line=tmp_line
        # diverse_num=0
        # for i in range(len(sort_score_dict_fusion)):
        #     diverse_num+=1
        #     tmp_row=row
        #     if sort_score_dict_fusion[i][1][-1] <0.5 and  label_freq[sort_score_dict_fusion[i][0]]['var'] >= 0.15 and len(sort_score_dict_fusion[i][0]) > 1 and diverse_num <= diverse_candidate_num:
        #         worksheet.cell(line,tmp_row,sort_score_dict_fusion[i][0]).fill=red_fill
        #     else:
        #         worksheet.cell(line,tmp_row,sort_score_dict_fusion[i][0])
                
        #     worksheet.cell(line,tmp_row,sort_score_dict_fusion[i][0])
        #     tmp_row+=1
        #     worksheet.cell(line,tmp_row,sort_score_dict_fusion[i][1][11])
        #     tmp_row+=1
        #     worksheet.cell(line,tmp_row,sort_score_dict_fusion[i][1][-1])
        #     tmp_row+=1
        #     worksheet.cell(line,tmp_row,label_freq[sort_score_dict_fusion[i][0]]['var'])
        #     line+=1
        #     tmp_row+=2
        # row=tmp_row

        # line=tmp_line
        # for i in range(len(sort_score_dict_fusion2)):
        #     tmp_row=row
        #     worksheet.cell(line,tmp_row,sort_score_dict_fusion2[i][0])
        #     tmp_row+=1
        #     worksheet.cell(line,tmp_row,sort_score_dict_fusion2[i][1][11])
        #     tmp_row+=1
        #     worksheet.cell(line,tmp_row,sort_score_dict_fusion2[i][1][-1])
        #     tmp_row+=1
        #     worksheet.cell(line,tmp_row,label_freq[sort_score_dict_fusion2[i][0]]['var'])
        #     line+=1
        # row=tmp_row

        
        line+=1
        # todo write in worksheet
    #todo save
    workbook.save(save_folder+object_excel)
    # save data_lst
    # with open('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Evaluate/yao_check_datas/annotation/0228for_annotation_train.json','w') as f:
    #     json.dump(new_data_lst,f)
        
    # 不同的预测范围的divserity
    # top`10
    # diversirty_top10=len(list(set(broaden_golden_labels_top10)-set(golden_labels)))
    # diversirty_top15=len(list(set(broaden_golden_labels_top15)-set(golden_labels)))
    diversity=len(list(set(broaden_labels)-set(golden_labels)))
    # print('f1:{}'.format(np.array(f1_lst).mean()))
    # # top`15
    # print('diversirty_top10:{},diversirty_top15:{}'.format(diversirty_top10,diversirty_top15))
    print('diversity:{}'.format(diversity))
    #todo save test_label_predict_infos
    # todo pearson correlation
    # pearson_rs=np.array(pearson_rs)
    # print('pearson correlation:{}'.format(pearson_rs.sum()/len(pearson_rs)))
    # with open('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Evaluate/yao_check_datas/train_song_infos.json','w') as f:
    #     json.dump(test_song_infos,f)

# test_data=json.load(open(test_f_path,'r'))
# train_data=json.load(open(train_f_path,'r'))
# print('all data len:{}'.format(len(test_data)+len(train_data)))
#todo 常用函数
# valid_score(all_vec_dict,vec_dict,test_f_path)


# 向文件中写入novelty score
def novelty_score(all_vec_dict,centers_vecs,data_lst):
    # 读取数据集
    vecs=list(all_vec_dict.values())
    # np to tensor
    for i in range(len(vecs)):
        vecs[i]=torch.tensor(vecs[i])
    vecs=torch.stack(vecs)
    # 计算所有vec的中心距离
    dists=mean_center_dist(vecs,centers_vecs)
    max_dist=max(dists)
    min_dist=min(dists)
    all_labels_dist_dict={}
    i=0
    for key in all_vec_dict:
        all_labels_dist_dict[key]=dists[i]
        i+=1
    #数据直接输入函数
    # with open(f_path,'r') as f:
    #     data_lst=json.load(f)
    i=0
    for data_item in data_lst:
        i+=1
        # 打印进度
        if i%100==0:
            print('processing {}/{} data'.format(i,len(data_lst)))
        # 获取所有candidate
        candidate_dict=data_item['final_score_dict_sort']
        for candiate in candidate_dict:
            candidate_dict[candiate][12]=1-all_labels_dist_dict[candiate]
        data_item['final_score_dict_sort']=candidate_dict
    # save data_lst
    # print('begin saving ...')
    # with open(save_path,'w') as f:
    #     json.dump(data_lst,f)
    # print('saved')
    return data_lst



#debug novelty score
centers_vec_origin='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/iter_labels/centers_iter-1_basedexpert.npy'
centers_vec_new='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/iter_labels/centers_iter1_0501_diva_finetune_base_expert.npy'
# centers=np.load(centers_vec_new,allow_pickle=True)
# data_lst=json.load(open('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/test_iter_based_expert.json','r'))
# novelty_score(all_vec_dict,centers,data_lst)


# todo 更新silver labels时同步更新novelty_score
def silver_update_noveltyS(centers_vecs,train_datas,test_datas,all_vec_dict=all_vec_dict):
    # centers_vec_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/iter_labels/centers_iter{}.npy'.format(iter)
    # centers_vecs=np.load(centers_vec_path,allow_pickle=True)
    # curent_f_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/eval_data/test_ebc_val_iter{}_{}(no_discard).json'.format(iter,date)
    # iter_train_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/train_iter{}.json'.format(iter)
    # iter_test_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/test_iter{}.json'.format(iter)
    test_datas=novelty_score(all_vec_dict,centers_vecs,test_datas)
    print('test novelty score updated')
    train_datas=novelty_score(all_vec_dict,centers_vecs,train_datas)
    print('train novelty score updated')
    return train_datas,test_datas


if __name__=='__main__':    
    # novelty_score(all_vec_dict,centers_vecs,origin_train_path,object_train_path)
    iter=0
    date='0330'
    # # centers_vec_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/iter_labels/centers_iter{}.npy'.format(iter_num-1)
    # # centers_vecs=np.load(centers_vec_path,allow_pickle=True)
    # curent_f_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/eval_data/test_ebc_val_iter{}_{}.json'.format(iter_num,date)
    # # iter_train_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/train_iter{}.json'.format(iter_num-1)
    # iter_test_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/test_iter{}.json'.format(iter_num-1)
    # # novelty_score(all_vec_dict,centers_vecs,iter_train_path,iter_train_path)
    # # novelty_score(all_vec_dict,centers_vecs,iter_test_path,iter_test_path)
    # novelty_score(all_vec_dict,centers_vecs,curent_f_path,curent_f_path)

    if iter>0:
        iter_train_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/train_iter{}_3.json'.format(iter-1)
        iter_test_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/test_iter{}_3.json'.format(iter-1)
    else:
        iter_test_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/test_t0_for_supervise_DIVA_0320.json'
    # # load iter2
    # # tmp_data=json.load(open(iter_test_path.format(2),'r'))
    iter_show_excel='iter{}_show_{}_3_095.xlsx'.format(iter,date)
    valid_score(all_vec_dict,vec_dict,iter_test_path,iter_show_excel)      
                    

                
        


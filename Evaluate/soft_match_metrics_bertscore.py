import  numpy as np
import torch 
import json
import re
import math
import openpyxl
#todo 获取bert score dict {song_name|candidate:bert_score}

# create excel
workbook=openpyxl.Workbook()
worksheet=workbook.active
worksheet.title='semantic soft match'

# todo bert score put into site -3


# some flags
USE_DICT=True

threshold=0.94

label_set='annotated+golden'

# file path
EmbeddingData_ROOT='/data/yuanxin_data/tme_big_data/song_id_candidate_words_embedding/'

#? data source
# candidate_w [song_id_xxxxx|w]
Candidates_FILE_train='train_song_id_candidate_words_shuffle_1108.npy'
Candidates_FILE_test='val_song_id_candidate_words_1108.txt'

# fasttext candidate embeddings `dict` type
# Candidates_dict_path='candidate_words_fasttext_embedding_dict.npy'
# Candidates_dict_path='candidate_words_CEM_embedding_dict.npy'
# epoch 10 candidate embeddings
# Candidates_dict_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/tme_big_data/vectors_epoch300.npy'
# Candidates_dict_path='/data/yuanxin_data/tme_big_data/song_id_candidate_words_embedding/candidate_words_context_embedding_dict_1536_195.npy'
Candidates_dict_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/tme_big_data/candidate_words_context_embedding_dict_1536_300.npy'

vec_dict=np.load(Candidates_dict_path,allow_pickle=True).item()
def get_bert_score(pred_label,true_labels,vec_dict=vec_dict):
    cos_sims=[]
    pred_vec=vec_dict[pred_label]
    for true_label in true_labels:
        true_vec=vec_dict[true_label]
        cos_sim=np.dot(pred_vec,true_vec)/(np.linalg.norm(pred_vec)*np.linalg.norm(true_vec))
        cos_sims.append(cos_sim)
    return max(cos_sims)

# embedding np.array
Embedding_FILE_train='train_song_candidate_words_embedding_shuffle_1108.npy'
Embedding_FILE_test='val_song_candidate_words_embedding_1108.npy'


# {song_id_xxxxx:vec}
song_id_comment_vec_path_train='train_song_id_layer_1_last_1108.json'
song_id_comment_vec_path_test='val_song_id_layer_1_last_1108.json'

device = torch.device("cuda",index=1 if torch.cuda.is_available() else "cpu")

# ��ʼ�ļ�song_info(train+test)
file_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/test_iter0_5.json'
new_file_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/pu_data/test_t0_for_supervise_ebc_1130_test39_puS_bleuS_bertS.json'

def load_data():
    test_lst = json.load( open(file_path))
    return test_lst

def load_npy_(npy_file_path):
    dict_data=np.load(npy_file_path,allow_pickle=True).item()
    # transform dict to l2index+candidate_embeddings
    l2index={}
    candidate_embeddings=[]
    index=0
    for item in dict_data:
        l2index[item]=index
        index+=1
        candidate_embeddings.append(dict_data[item])
    return l2index,candidate_embeddings

#f_path1 记录所有的candidate_w(txt或npy) f_path2记录所有的candidate_w_vec(npy)
# f_path1,2 是train f_path3,4是test
def load_embeddings(f_path0,f_path1):
    if '.npy' in f_path0:
        candidates=np.load(f_path0)
    else:
        # for txt file
        f = open(f_path0,encoding = "utf-8")
        candidates=f.read()
        candidates=candidates.split('\n')
    l2index=get_word2index(candidates)
    candidates_embedding=np.load(f_path1)
    return l2index,candidates_embedding

#todo merge label_embeddings
def embeddings_merge(l2index_train,l2index_test,candidates_embedding_train,candidates_embedding_test):
    l2index={}
    candidates_embedding=[]
    # 先导入train
    index=0
    for l_info in l2index_train:
        if l_info not in l2index:
            l2index[l_info]=len(l2index)
            candidates_embedding.append(candidates_embedding_train[index])
            index+=1
    # 再导入test
    index=0
    for l_info in l2index_test:
        if l_info not in l2index:
            l2index[l_info]=len(l2index)
            candidates_embedding.append(candidates_embedding_test[index])
            index+=1
    return l2index,candidates_embedding
    
# 构造word2index 便于根据candidatew 找到对应向量
def get_word2index(candidate_w_array):
    # l2index{label_info(song_id,label):vec}
    l2index={}
    index=0
    for l_info in candidate_w_array:
        l_info=re.sub('\n','',l_info)
        l2index[l_info]=index
        index+=1
    return l2index


def greedy_cos_idf_labels_term(
    ref_embedding,
    hyp_embedding,
    threshold
):
    """
        Compute greedy matching based on cosine similarity.

        Args:
            - :param: `ref_embedding` (torch.Tensor):
                    embeddings of reference sentences, BxKxd,
                    B: batch size, K: longest length, d: bert dimenison
                    labels_embeding:[1,labels_num,768*2]
            - :param: `ref_lens` (list of int): list of reference sentence length. labels_num
            - :param: `hyp_embedding` (torch.Tensor):
                    embeddings of candidate sentences, BxKxd,
                    B: batch size, K: longest length, d: bert dimenison
                    candidate_w embedding [1,candidate_num,768*2]
            - :param: `hyp_lens` (list of int): list of candidate sentence length.
    """
    ref_embedding.div_(torch.norm(ref_embedding, dim=-1).unsqueeze(-1))
    hyp_embedding.div_(torch.norm(hyp_embedding, dim=-1).unsqueeze(-1))
    sim = torch.bmm(hyp_embedding, ref_embedding.transpose(1, 2))
    word_recall=torch.max(sim[0],dim=1).values.to(torch.device('cpu')).numpy()
    # word_recall_debug=torch.max(sim[0],dim=0)
    # word_recall_debug2=torch.max(sim[0],dim=1)
    # word_recall = sim.max(dim=1)[0][0]
    # max_sim=torch.max(word_recall)
    return word_recall


def semantic_soft_match(song_id,pred_ls,golden_ls,l2index,label_embedings,threshold,device):
    """
        Compute y_pred based on semantic soft matching.

        Args:
            - :param: `pred_l` (list):
            - :param: `golden_l` (list of int): list of reference sentence length. labels_num
            - :param: `w2embeding` (dict:{w:w_vec}): w_vec:[768*2]tensor
            - :param: `threshold` (float).
            - :param: `device` (torch.device): cuda // cpu.
        return:
            - :dict: `candidate_w : bert_score * 1/rank`
        
    """
    # 输出 y_true
    # 将golden_ls转换成[1*num*(728*2)]形式�?? 构建ref_embedding
    y_pred=[]
    golden_embds=[]
    # candidate : bert_score
    bert_score_dict={}
    # candidate : rank
    bert_score_rank={}
    for golden_l in golden_ls:
        if not USE_DICT:
            key_tmp=song_id+'|'+golden_l
        else:
            key_tmp=golden_l
        try:
            label_index=l2index[key_tmp]
        except KeyError as Exception:
            print('golden no embeddings:{}'.format(key_tmp))
            continue
        label_vec=label_embedings[label_index]
        golden_embds.append(torch.tensor(label_vec).to(device))
    
    if len(golden_embds)==0:
        return [0]*len(pred_ls)
    ref_embds=torch.stack(golden_embds)
    ref_embds=torch.stack([ref_embds])

    # 构建pred_embedding [1*1*(768*2)]
    pred_embds=[]
    for pred_l in pred_ls:
        try:
            if USE_DICT:
                pred_l_index=l2index[pred_l]
            else:
                pred_l_index=l2index[song_id+'|'+pred_l]
        except KeyError as Exception:
            print('predict no embeddings:{}'.format(song_id+'|'+pred_l))
            continue
        pred_vec=torch.tensor(label_embedings[pred_l_index]).to(device)
        pred_embds.append(pred_vec)
        # input pre
    hyp_embds=torch.stack(pred_embds)
    hyp_embds=torch.stack([hyp_embds])
    bert_scores=greedy_cos_idf_labels_term(ref_embds,hyp_embds,threshold)
    index=0
    for candiadate in pred_ls:
        if USE_DICT:
            candiadate_info=candiadate
        else:
            candiadate_info=song_id+'|'+candiadate
        if candiadate_info not in l2index:
            bert_score_dict[candiadate]=0
            continue
        bert_score_dict[candiadate]=bert_scores[index]
        index+=1
    # get bert score rank
    bertscore_sort=sorted(bert_score_dict.items(),key=lambda val:val[1],reverse=True)
    last_bert_score=1
    index=0
    tmp_rank=1
    for candidate_info in bertscore_sort:
        index+=1
        if candidate_info[1]< last_bert_score:
            bert_score_rank[candidate_info[0]]=math.log((index+1),2)
            tmp_rank=index
        else:
            bert_score_rank[candidate_info[0]]=math.log((tmp_rank+1),2)
    
    return bert_score_dict,bert_score_rank

def norm_song_id(song_name):
    pat=re.compile('song_id: (.*?) ')
    song_id=pat.findall(song_name)[0]
    return 'song_id_'+song_id.strip()

def main():
    print('loading data...')
    saved_folder_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Evaluate/soft_match_score/'
    song_info_dict=json.load(open('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Evaluate/yao_check_datas/test_song_infos.json','r'))
    file_name='bert_score_dict({}).json'.format(label_set)
    bert_score_dict={}
    data_lst=load_data()
    if not USE_DICT:
        l2index_train,label_embeddings_train=load_embeddings(EmbeddingData_ROOT+Candidates_FILE_train,EmbeddingData_ROOT+Embedding_FILE_train)
        l2index_test,label_embeddings_test=load_embeddings(EmbeddingData_ROOT+Candidates_FILE_test,EmbeddingData_ROOT+Embedding_FILE_test)
        # 合并train test embeddings
        l2index,label_embeddings=embeddings_merge(l2index_train,l2index_test,label_embeddings_train,label_embeddings_test)
    else:
        # l2index,label_embeddings=load_npy_(EmbeddingData_ROOT+Candidates_dict_path)
        l2index,label_embeddings=load_npy_(Candidates_dict_path)
    for song_comment in data_lst:
        song_name=song_comment['song_name']
        if song_name not in song_info_dict:
            continue
        song_labels_info=song_info_dict[song_name]
        song_id=norm_song_id(song_name)
        if label_set=='golden':
            label_lst=song_comment['song_pseudo_golden_labels']
        elif label_set=='expert':
            label_lst=song_labels_info['expert_labels']
        elif label_set=='annotated+golden':
            label_lst=song_labels_info['annotated_golden_labels']
        else:
            label_lst=song_labels_info['annotated_expert_labels']
        final_score_dict=song_comment['final_score_dict_sort']
        candidate_lst=list(final_score_dict.keys())
        bert_scores_dict,bert_scores_rank=semantic_soft_match(song_id,candidate_lst,label_lst,l2index,label_embeddings,threshold,device)
        for candidate in final_score_dict:
            # add rank
            # final_score_dict[candidate][-3]=bert_scores_dict[candidate]/bert_scores_rank[candidate]
            # final_score_dict[candidate][-3]=float(round(bert_scores_dict[candidate],5))
            tmp_bertS=float(round(bert_scores_dict[candidate],5))
            bert_score_dict['|'.join([song_name,candidate])]=tmp_bertS
            
            
            # print(type(round(bert_scores_dict[candidate],3)))
        # save all semantice soft match score
        # bert_score_dict_sort=sorted(bert_scores_dict.items(),key=lambda val:val[1],reverse=True)
        # for item in bert_score_dict_sort:
        #     worksheet.cell(line,1,item[0])
        #     worksheet.cell(line,2,item[1])
        #     line+=1
    
    # save workbook
    # workbook.save('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Evaluate/yao_check_datas/semantic_soft_match.xlsx')

    # save new json
    print('json file is writing...')
    # with open(new_file_path,'w') as f:
    #     json.dump(data_lst,f,ensure_ascii=False)
    # save bert score dict
    with open(saved_folder_path+file_name,'w') as f:
        json.dump(bert_score_dict,f,ensure_ascii=False)
            
        
if __name__=='__main__':
    main()
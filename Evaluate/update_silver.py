import json
import numpy as np
import math
import openpyxl
import os

import sys
sys.path.append("..")
from Evaluate.k_means_yao import M_N_clustring
from Evaluate.for_valid_score import silver_update_noveltyS

# from k_means_yao import M_N_clustring
# from for_valid_score import silver_update_noveltyS
from tkinter import _flatten
import xlsxwriter
import re
import torch

from line_profiler import LineProfiler



device=torch.device('cuda' if torch.cuda.is_available() else 'cpu')

# whether a str is num

def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        pass
    try:
        import unicodedata
        unicodedata.numeric(s)
        return True
    except (TypeError, ValueError):
        pass
    return False


#todo 词频字典
#todo valid score - 离散系数
def cal_valid_score():
    # label : valid score
    label_freq=json.load(open('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Evaluate/yao_check_datas/words_tf_info_freq.json'))
    max_val=0
    min_val=1
    for key in label_freq.keys():
        tf_infos=np.array(label_freq[key]['freq'])
        # valid_score=np.var(tf_infos)
        valid_score=np.std(tf_infos)/np.mean(tf_infos)
        label_freq[key]['var']=valid_score
        if valid_score>max_val:
            max_val=valid_score
        if valid_score<min_val:
            min_val=valid_score
    # max-min normalization
    for key in label_freq.keys():
        label_freq[key]['var']=(label_freq[key]['var']-min_val)/(max_val-min_val)

    return label_freq
label_freq=cal_valid_score()

def get_annotated_info(f_path):
    song_info_dict=json.load(open('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Evaluate/yao_check_datas/test_song_infos.json','r'))
    data_lst=json.load(open(f_path,'r'))
    for item in data_lst:
        song_id=item['song_name']
        if song_id in song_info_dict:
            item['song_annotated_labels']=song_info_dict[song_id]['annotated_labels']
    # save json
    json.dump(data_lst,open(f_path,'w'))
    print('annotated labels writed')

# get_annotated_info(silver_iter1_test)



def get_valid_label(f_path,ebc_val_f_path,N=16372):
    # golden label 频率字典
    label_freq=json.load(open('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Evaluate/yao_check_datas/words_tf_info_freq.json'))
    data_lst=json.load(open(f_path,'r'))
    ebc_val_lst=json.load(open(ebc_val_f_path,'r'))
    for i in range(len(data_lst)):
        if i%100==0:
            print(i)
        item=data_lst[i]
        score_dict=item['final_score_dict_sort']
        score_dict_ebc=ebc_val_lst[i]['final_score_dict_sort']
        for key in score_dict:
            tf_infos=np.array(label_freq[key]['freq'])
            dn=score_dict[key][2]
            tn_mean=tf_infos.mean()
            tn_max=tf_infos.max()
            valid_score=(tn_max-tn_mean)/min(0.5*math.log(dn+1)*math.log(N/(dn+1)),math.log(dn+1))
            score_dict[key][-2]=valid_score
            score_dict[key][-1]=score_dict_ebc[key][-1]
        item['final_score_dict_sort']=score_dict
    # save json
    json.dump(data_lst,open(f_path,'w'))
    print('valid score writed')

# get_valid_label(test_f_path,ebc_test_f_path)


def max_min_normalization(test_f_path,train_f_path,object_test_f_path,object_train_f_path):
    max_val=0
    min_val=1
    test_lst=json.load(open(test_f_path,'r'))
    train_lst=json.load(open(train_f_path,'r'))
    print('file loaded')
    for item in test_lst+train_lst:
        score_dict=item['final_score_dict_sort']
        for key in score_dict:
            if score_dict[key][12]>max_val:
                max_val=score_dict[key][12]
            if score_dict[key][12]<min_val:
                min_val=score_dict[key][12]
    print(max_val,min_val)
    for item in test_lst:
        score_dict=item['final_score_dict_sort']
        for key in score_dict:
            score_dict[key][12]=(score_dict[key][12]-min_val)/(max_val-min_val)
        item['final_score_dict_sort']=score_dict
    print('test data update done')
    for item in train_lst:
        score_dict=item['final_score_dict_sort']
        for key in score_dict:
            score_dict[key][12]=(score_dict[key][12]-min_val)/(max_val-min_val)
        item['final_score_dict_sort']=score_dict
    print('train data update done')
    json.dump(test_lst,open(object_test_f_path,'w'))
    json.dump(train_lst,open(object_train_f_path,'w'))

# max_min_normalization(test_f_path,train_f_path,object_test_f_path,object_train_f_path)
# 更新silver label
vec_path='DiVa/data/tme_big_data/golden_dict_1536_300_pca512.npy'
# golden_dict_xlnet.npy
# vec_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/tme_big_data/golden_dict_xlnet.npy'
vec_dict=np.load(vec_path,allow_pickle=True).item()
# 所有golden labels 

# todo get harmonic mean
def harm_mean(lst,p=1e-8):
    sum=0
    for item in lst:
        sum+=1/(item+p)
    return 1/sum

# todo 用来查看扩增的golden labels
golden_labels=list(vec_dict.keys())
song_candidate_words_dict = np.load('DiVa/data/tme_big_data/candidate_words_context_embedding_dict_1536_300_pca512.npy', allow_pickle=True).item()
all_available_labels=list(song_candidate_words_dict.keys())
def update_silver_iter1(data_lst,save_f_path,iter,ebc_s_dict,valid_score_dict,with_novelty,with_tf_idf,with_valid,PE_thres,ebc_thres,joint_score_thres,label_class='expert',all_goldens=golden_labels,label_freq=label_freq,save=False,uncertainty_limit=False,tf_idf_type='old',write_in_xlsx=False,update_thres=0.05,dumping_weight=0.5,with_valid_semantic=True,with_valid_disc=True,only_joint=False,write_file_name='demo'):
    """
    data_lst: train / test data list
    save_f_path : updated data save path
    iter : int
    tf_idf_type: str:: new site 4/ old site 10
    with_novelty: bool wether use novelty score to generate silver labels
    with_tf_idf: bool wether use tf_idf score to generate silver labels
    with_valid: bool wether use valid score to generate silver labels
    PE_thres: float:: the threshold of PE score to generate silver labels
    ebc_thresh: float:: the threshold of ebc score to generate silver labels
    update_thres: choose how many candidate to update tp silver -- cal silver update num
    dumping_weight: float:: the weight of dumping score(& update silver num)
    write_file_name: str:: the name of excel file(if write_in_xlsx==True)
    """
    # if iter==1:
    #     update_discard_rate=1
    # else:
    update_discard_rate=math.pow(dumping_weight,iter)
    print('update_discard_rate',update_discard_rate)
    print('with_novelty',with_novelty)
    print('with_tf_idf',with_tf_idf)
    print('with_valid_semantic',with_valid_semantic)
    print('with_valid_disc',with_valid_disc)

    if write_in_xlsx:
        workbook=openpyxl.Workbook()
        worksheet=workbook.active
        row=1
    # save json
    # scores_save_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Evaluate/yao_check_datas/data_distribution_imgs/iter0_test_dict.json'
    iter_scores_dict={}
    line=1
    # 记录100首train data的silver label变化
    all_labels=[]
    all_labels_pred=[]
    # data_lst=json.load(open(f_path,'r'))
    j=0
    diverse_score_site=0
    write_excel=0
    print(len(data_lst))
    # train 集合打印silver label变化
    if len(data_lst)>1000:
        write_excel=1
    for item in data_lst:
        j+=1
        if j%100==0:
            print(j)
            
        song_name=item['song_name']
        score_dict=item['final_score_dict_sort']
        silver_update_num=update_thres*len(score_dict)
        if label_class=='golden':
            golden_labels=item['song_pseudo_golden_labels']
        else:
            golden_labels=list(set(item['song_labels']) & set(all_available_labels))
        all_silvers=[]
        # 综合考虑golden标签数量以及candidate数量来扩增silver labels
        golden_num=len(golden_labels)
        if 'song_all_silver_labels' in item:
            all_silvers=item['song_all_silver_labels']
        # 用于计算PE score all_labels_1所有应该被预测为1的labeL
        all_labels_1=list(set(golden_labels+all_silvers))
        score_dict_sort=sorted(score_dict.items(),key=lambda x:x[1][-1],reverse=True)
        candidate_num=len(score_dict_sort)
        silver_num=min((candidate_num*update_thres+golden_num)/2,candidate_num*update_thres)
        if write_in_xlsx and j<=50:
            worksheet.cell(row=row,column=1,value=song_name)
            worksheet.cell(row=row,column=2,value=silver_num)
            row+=1
            worksheet.cell(row=row,column=1,value='labels')
            worksheet.cell(row=row,column=2,value='fusion score')
            # worksheet.cell(row=row,column=3,value='valid-1')
            # worksheet.cell(row=row,column=4,value='valid-2')
            worksheet.cell(row=row,column=3,value='valid')
            worksheet.cell(row=row,column=4,value='tf-idf')
            worksheet.cell(row=row,column=5,value='novelty')
            worksheet.cell(row=row,column=6,value='PE')
            worksheet.cell(row=row,column=7,value='ebc')
            
            # worksheet.cell(row=row,column=4,value='novd score')
            row+=1
                    
        song_labels=list(set(golden_labels+all_silvers))
        # sort by site -1
        # get candidate num
        # get rank by site -1
        rank_dict={}
        max_val=1
        ebc_more_than_095=0
        # ebc 大于0.95 的silver
        ebc_lst=[]
        ebc_pred_lst=[]
        ebc_rank=0
        # hard_sample_extent_dict={}
        label_uncertainty_infos={}
        
        labels_pred_more_than_095=[]
        for i in range(len(score_dict_sort)):
            # if i < silver_update_num:
            #     ebc_pred_lst.append(score_dict_sort[i][0])
            #     if i < silver_num * update_discard_rate:
            #         ebc_lst.append(score_dict_sort[i][0])
            # else:
            #     break
            # if score_dict_sort[i][1][-1] < max_val:
            #     max_val=score_dict_sort[i][1][-1]
            #     ebc_rank=i+1
            if score_dict_sort[i][1][-1] >= ebc_thres:
                ebc_more_than_095+=1
                ebc_lst.append(score_dict_sort[i][0])
                ebc_pred_lst.append(score_dict_sort[i][0])
            # else:
            #     break
            # rank_dict[score_dict_sort[i][0]]=ebc_rank
            
            #todo 评估模型是否预测新标签，用0.95作为阈值
            # if score_dict_sort[i][1][-1] >= 0.95:
            #     labels_pred_more_than_095.append(score_dict_sort[i][0])
            #todo entropy_uncertainty
            if uncertainty_limit:
                # get PE_score
                if score_dict_sort[i][0] in all_labels_1:
                    label_uncertainty=-(score_dict[score_dict_sort[i][0]][-1]*math.log(score_dict[score_dict_sort[i][0]][-1]))
                else:
                    label_uncertainty=-((1-score_dict[score_dict_sort[i][0]][-1])*math.log((1-score_dict[score_dict_sort[i][0]][-1])))
                label_uncertainty_infos[score_dict_sort[i][0]]=label_uncertainty
        
        # new_labels generate by model 
        all_labels_pred.extend(ebc_pred_lst)
        
        #todo 如果方法可行->换成中位数来判断是否处于�?50%
        # sort hard sample extent
        if uncertainty_limit:
            label_uncertainty_dict_sort=sorted(label_uncertainty_infos.items(),key=lambda x:x[1],reverse=True)
            
            # get hard sample extent rank
            label_uncertainty_rank_dict={}
            for i in range(len(label_uncertainty_dict_sort)):
                label_uncertainty_rank_dict[label_uncertainty_dict_sort[i][0]]=(i+1)/candidate_num
        
        #v silver num = mean(candiate num 8 0.05,ebc_more_than_095)
        # silver_num=min(int((candidate_num*0.05+ebc_more_than_095)/2),int(candidate_num*0.05))

        # label: [match , valid , diverse (novelty with tf-idf++) , hard_sample_extent]
        silver_dict={}
        for key in score_dict:
            #todo match score
            match_score=score_dict[key][-1]
            #todo valid score
            # valid_score=valid_score_dict[key]['valid_score']
            #todo diverse_score (novelty with tf-idf++) 
            # novelty_match=math.pow(score_dict[key][12],2)+math.pow(score_dict[key][-1],2)
            if tf_idf_type=='new':
                tf_idf_s=score_dict[key][4]
            else:
                tf_idf_s=score_dict[key][10]
                
            novelty_s=score_dict[key][12]
            
            valid_semantic=valid_score_dict[key]['valid_semantic']
            valid_disc=valid_score_dict[key]['valid_disc']
            valid_s=valid_score_dict[key]['valid_score']
            lenth_pen=valid_score_dict[key]['length_punish']
            #todo valid score
            # valid_s=lenth_pen*harm_mean([valid_semantic,valid_disc])
            # valid_s=lenth_pen*valid_semantic*valid_disc
            
            # if iter ==0:
            #     valid_sema_thres=0.01
            # else:
            #     valid_sema_thres=0.01 *  min(10*(iter),50)
            # if valid_semantic >= valid_sema_thres:
            #     valid_semantic=1
            # else:
            #     valid_semantic=0
                    
            # if valid_disc>=0.1:
            #     valid_disc=1
            # else:
            #     valid_disc=0
            # valid_s=1/(1/(valid_semantic+1e-8)+1/(valid_disc+1e-8))
            # valid_s=math.log(valid_s+1,2)
            # diverse_sore=math.pow(tf_idf_s*valid_s,0.5)
            if with_tf_idf==False:
                tf_idf_s=novelty_s
                # 这样求调和平均就是其本身
            if with_novelty==False:
                novelty_s=tf_idf_s
            if with_valid_semantic==False:
                valid_semantic=0
                # valid_s=lenth_pen*valid_disc
            if with_valid_disc==False:
                valid_disc=0
            valid_s=(valid_disc+valid_semantic)*lenth_pen 
                
            final_s=valid_semantic* valid_disc * tf_idf_s * novelty_s
      
            # diverse_score_site+=1
            #todo hard_sample_extent
            if uncertainty_limit:
                hard_sample_extent=label_uncertainty_rank_dict[key]
            else:
                hard_sample_extent=0
            # freq_match=score_dict[key][4]/math.sqrt(rank_dict[key])
            # valid_score=label_freq[key]['var']
            # silver_dict[key]=[match_score,novelty_match,freq_match,valid_score]
            # silver_dict[key]=[match_score,math.sqrt(tf_idf_s*novelty_s),valid_s,tf_idf_s,novelty_s,hard_sample_extent]
            silver_dict[key]=[match_score,harm_mean([tf_idf_s,novelty_s]),valid_s,tf_idf_s,novelty_s,hard_sample_extent]
            # silver_dict[key]=[match_score,final_s,hard_sample_extent]
        
        # sort by 1
        # valid_score_rank_dict={}
        # for i in range(len(silver_dict_sort_valid)):
        #     valid_score_rank_dict[silver_dict_sort_valid[i][0]]=i+1
        
        # # rewrite silver_dict (diverse_score)
        # for key in silver_dict:
        #     silver_dict[key][2]/=math.log(valid_score_rank_dict[key]+1,2)
        sort_dict_fusion=sorted(silver_dict.items(),key=lambda x:x[1][2],reverse=True)
        #todo by rank
        valid_rank_dict={}
        if with_valid_disc==False and with_valid_semantic==False:
            for i in range(len(sort_dict_fusion)):
                valid_rank_dict[sort_dict_fusion[i][0]]=1
        else:
            for i in range(len(sort_dict_fusion)):
                valid_rank_dict[sort_dict_fusion[i][0]]=i+1
            
        for l in silver_dict:
            # silver_dict[l][1]/=math.log(valid_rank_dict[l]+1,2)
            # silver_dict[l][1]/=math.sqrt(valid_rank_dict[l])
            silver_dict[l][1]*=(1-valid_rank_dict[l]/candidate_num)
            
        sort_dict_final=sorted(silver_dict.items(),key=lambda x:x[1][1],reverse=True)
        #todo by rank

        if write_in_xlsx and j<=50:
            for i in range(len(sort_dict_fusion)):
            #       worksheet.cell(row=row,column=1,value=song_name)
            # row+=1
            # worksheet.cell(row=row,column=1,value='labels')
            # worksheet.cell(row=row,column=2,value='fusion score')
            # # worksheet.cell(row=row,column=3,value='valid-1')
            # # worksheet.cell(row=row,column=4,value='valid-2')
            # worksheet.cell(row=row,column=3,value='valid')
            # worksheet.cell(row=row,column=4,value='tf-idf')
            # worksheet.cell(row=row,column=5,value='novelty')
            # worksheet.cell(row=row,column=6,value='PE')
            # worksheet.cell(row=row,column=7,value='ebc')
                worksheet.cell(row=row,column=1,value=sort_dict_final[i][0])
                worksheet.cell(row=row,column=2,value=sort_dict_final[i][1][1])
                worksheet.cell(row=row,column=3,value=sort_dict_final[i][1][2])
                worksheet.cell(row=row,column=4,value=sort_dict_final[i][1][3])
                worksheet.cell(row=row,column=5,value=sort_dict_final[i][1][4])
                worksheet.cell(row=row,column=6,value=sort_dict_final[i][1][-1])
                worksheet.cell(row=row,column=7,value=sort_dict_final[i][1][0])
                row+=1
        # ebc_s 0.05~0.0001 valid_score >= 0.001
        # novelty_match_lst=[]
        # for i in range(len(silver_dict_sort_site1)):
        #     if silver_dict_sort_site1[i][1][1] >= 1:
        #         if silver_dict_sort_site1[i][1][0] <0.05 and silver_dict_sort_site1[i][1][0] > 0.0001 and silver_dict_sort_site1[i][1][3] >= 0.14 and is_number(silver_dict_sort_site1[i][0])==False and len(silver_dict_sort_site1[i][0])>1:
        #             novelty_match_lst.append(silver_dict_sort_site1[i][0])
        #     else:
        #         break
 
        # all_silver_up_013.extend(silver_labels_count_013)
        
        diverse_lst=[]
        for i in range(len(sort_dict_final)):
            if uncertainty_limit==True:
                if i < silver_num * update_discard_rate:
                    if sort_dict_final[i][1][-1] < PE_thres and is_number(sort_dict_final[i][0])==False:
                        diverse_lst.append(sort_dict_final[i][0])
                else:
                    break
            else:
                if i < silver_num * update_discard_rate:
                    if sort_dict_final[i][1][1]<=0:
                        break
                    if is_number(sort_dict_final[i][0])==False:
                        diverse_lst.append(sort_dict_final[i][0])
                else:
                    break
        
        if only_joint==False:
            if iter>=1:
                item['song_silver_labels'].append(list(set(diverse_lst+ebc_lst)))
            else:
                item['song_silver_labels']=[list(set(diverse_lst+ebc_lst))]
        else:
            if iter>=1:
                item['song_silver_labels'].append(list(set(diverse_lst)))
            else:
                item['song_silver_labels']=[list(set(diverse_lst))]
        
        #todo 删除 song_silver_labels中一次迭代还未学到内容的标签
        # new_song_silver_labels=[]
        # if 'song_silver_labels' in item:
        #     for lst in item['song_silver_labels']:
        #         tmp_song_silver_labels=[]
        #         for label in lst:
        #             if silver_dict[label][3] > 0.1:
        #                 tmp_song_silver_labels.append(label)
        #         new_song_silver_labels.append(tmp_song_silver_labels)
        #     item['song_silver_labels']=new_song_silver_labels

        # item['song_silver_labels']=[list(set(novelty_match_lst+freq_match_lst+ebc_lst))]
        item['song_all_silver_labels']=list(set(_flatten(item['song_silver_labels'])))
        # all_labels+=(item['song_silver_labels']+golden_labels)
        all_labels+=golden_labels
        for lst in item['song_silver_labels']:
            all_labels+=lst
        # all_labels+=item['song_all_silver_labels']
        
        # if i<=50 and write_excel==1:
        #     worksheet.cell(line,1,item['song_name'])
        #     line+=1
        #     worksheet.cell(line,1,'candidate_num')
        #     worksheet.cell(line,2,candidate_num)
        #     line+=1
        #     worksheet.cell(line,1,'golden')
        #     worksheet.cell(line,2,' '.join(golden_labels))
        #     line+=1
        #     all_silver=[]
        #     for i in range(len(item['song_silver_labels'])):
        #         all_silver+=item['song_silver_labels'][i]
        #         all_silver=list(set(all_silver))
        #         worksheet.cell(line,1,'silver'+str(i))
        #         worksheet.cell(line,2,' '.join(all_silver))
        #         line+=1
    # if write_excel==1:
    #     if label_class=='golden':
    #         workbook.save('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Evaluate/yao_check_datas/iter_check/silver_labels_to_iter{}_6.xlsx'.format(iter))
    #     if label_class=='expert':
    #         workbook.save('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Evaluate/yao_check_datas/iter_check/silver_labels_to_iter{}_5expert.xlsx'.format(iter))
            

    # save to file
    # diversity
    # diversity=len(set(all_labels)-set(all_goldens))
    # print('diversity',diversity)
    # save json
    # save to json (iter0 scores dict)
    # json.dump(iter_scores_dict,open(scores_save_path,'w'))
    # print('save json...')
    # all_tags_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/iter_labels/all_tags.json'
    # laod all tags
    # print('new_labels_up_13',len(new_labels_up_13))
    
    if write_in_xlsx:
        workbook.save('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Evaluate/yao_check_datas/{}.xlsx'.format(write_file_name))
    if save:
        print('save json')
        json.dump(data_lst,open(save_f_path,'w'))
        return all_labels_pred
    else:
        return all_labels,all_labels_pred,data_lst

#todo 获取ebc的分数分�?
def get_ebc_score_distribution(data_lst):
    """
    input: data_lst
    output: score_dict
    """
    ebc_s_dict={}
    for item in data_lst:
        score_dict=item['final_score_dict_sort']
        for key in score_dict:
            if key not in ebc_s_dict:
                ebc_s_dict[key]=[]
            ebc_s_dict[key].append(score_dict[key][-1])
    return ebc_s_dict

def get_mean(lst):
    val=0
    for item in lst:
        val+=0
    print(len(lst))
    return val/len(lst)

def get_valid_score_distribution(data_lst,ebc_s_dict,label_freq):
    """
    input: data_lst,ebc_s_dict
    output: valid_score_dict
    """
    valid_score_dict={}
    i=0
    for item in data_lst:
        i+=1
        if i%100==0:
            print('get_valid_score_distribution--{}'.format(i))
        score_dict=item['final_score_dict_sort']
        for key in score_dict.keys():
            # hhh=0
            if key not in valid_score_dict:
                valid_score_dict[key]={}
            discrete_coff=label_freq[key]['var']
            mean_ebc_s=np.array(ebc_s_dict[key]).mean()
            length_punish=math.exp(min(0,1-2/len(key)))
            length_punish=1
            valid_score_semantic=length_punish*mean_ebc_s
            valid_score_disc=length_punish*discrete_coff
            valid_score_dict[key]['discrete_coff']=discrete_coff
            valid_score_dict[key]['mean_ebc_s']=mean_ebc_s
            valid_score_dict[key]['length_punish']=length_punish
            valid_score_dict[key]['valid_semantic']=valid_score_semantic
            valid_score_dict[key]['valid_disc']=valid_score_disc
            valid_score_dict[key]['valid_score']=(valid_score_semantic+valid_score_disc)*length_punish
    return valid_score_dict

#todo 计算site4 * site 12
def cal_diverse_score(train_data,test_data,device=device):
    """
    input: train_data,test_data
    output: diverse_score_lst_train,diverse_score_lst_test
    """
    tf_idf_vals=[]
    novelty_vals=[]
    diverse_score_lst_train=[]
    diverse_score_lst_test=[]
    #test
    item_site=0
    for item in test_data:
        item_site+=1
        score_dict=item['final_score_dict_sort']
        for key in score_dict:
            tf_idf_vals.append(score_dict[key][4])
            novelty_vals.append(score_dict[key][12])
        if len(tf_idf_vals)>=100000:
            # list to tensor
            tf_idf_vals=torch.tensor(tf_idf_vals).to(device)
            novelty_vals=torch.tensor(novelty_vals).to(device)
            # cal score
            diverse_score_lst_test_tmp=tf_idf_vals*novelty_vals
            # tensor to list(cpu)
            diverse_score_lst_test+=diverse_score_lst_test_tmp.cpu().tolist()
            tf_idf_vals=[]
            novelty_vals=[]
        elif item_site==len(test_data):
            # list to tensor
            tf_idf_vals=torch.tensor(tf_idf_vals).to(device)
            novelty_vals=torch.tensor(novelty_vals).to(device)
            # cal score
            diverse_score_lst_test_tmp=tf_idf_vals*novelty_vals
            # tensor to list(cpu)
            diverse_score_lst_test+=diverse_score_lst_test_tmp.cpu().tolist()
            tf_idf_vals=[]
            novelty_vals=[]
    # list to tensor
    tf_idf_vals=torch.tensor(tf_idf_vals).to(device)
    novelty_vals=torch.tensor(novelty_vals).to(device)
    # cal score
    diverse_score_lst_test=tf_idf_vals*novelty_vals
    # tensor to list(cpu)
    diverse_score_lst_test=diverse_score_lst_test.cpu().tolist()

    tf_idf_vals=[]
    novelty_vals=[]
    # train 
    item_site=0
    for item in train_data:
        item_site+=1
        score_dict=item['final_score_dict_sort']
        key_site=0
        for key in score_dict:
            key_site+=1
            tf_idf_vals.append(score_dict[key][4])
            novelty_vals.append(score_dict[key][12])
        if len(tf_idf_vals)>=100000:
            # list to tensor
            tf_idf_vals=torch.tensor(tf_idf_vals).to(device)
            novelty_vals=torch.tensor(novelty_vals).to(device)
            # cal score
            diverse_score_lst_train_tmp=tf_idf_vals*novelty_vals
            # tensor to list(cpu)
            diverse_score_lst_train+=diverse_score_lst_train_tmp.cpu().tolist()
            tf_idf_vals=[]
            novelty_vals=[]
        elif item_site==len(train_data):
            # list to tensor
            tf_idf_vals=torch.tensor(tf_idf_vals).to(device)
            novelty_vals=torch.tensor(novelty_vals).to(device)
            # cal score
            diverse_score_lst_train_tmp=tf_idf_vals*novelty_vals
            # tensor to list(cpu)
            diverse_score_lst_train+=diverse_score_lst_train_tmp.cpu().tolist()
            tf_idf_vals=[]
            novelty_vals=[]
        
    return diverse_score_lst_train,diverse_score_lst_test
    

def update_silver(iter,all_goldens=golden_labels,object_test_f_path=object_test_f_path,object_train_f_path=object_train_f_path):
    # more_silver=0
    silver_iter_test='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/test_iter{}_5.json'
    silver_iter_train='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/train_iter{}_5.json'
    all_labels_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/iter_labels/all_labels_iter{}_5.json'.format(iter)
    if iter==0:
        test_datas=json.load(open(object_test_f_path))
        train_datas=json.load(open(object_train_f_path))
    else:
        test_datas=json.load(open(silver_iter_test.format(iter-1)))
        train_datas=json.load(open(silver_iter_train.format(iter-1)))
    print('load data over')
    
    # # train_diverse_score_lst,test_diverse_score_lst=cal_diverse_score(train_datas,test_datas)
    # print('diverse score analys over')
    ebc_s_dict=get_ebc_score_distribution(train_datas+test_datas)
    print('ebc score distribution analys over')
    valid_score_dict=get_valid_score_distribution(train_datas+test_datas,ebc_s_dict,label_freq)
    print('valid score distribution analys over')
    


    all_test_labels=update_silver_iter1(test_datas,silver_iter_test.format(iter),iter,ebc_s_dict,valid_score_dict,uncertainty_limit=False,save=False,label_class='golden')
    print('update silver iter{} test done'.format(iter))
    all_train_labels=update_silver_iter1(train_datas,silver_iter_train.format(iter),iter,ebc_s_dict,valid_score_dict,uncertainty_limit=False,save=False,label_class='golden')
    print('update silver iter{} train done'.format(iter))

    all_labels=list(set(all_test_labels+all_train_labels))
    # save all labels
    json.dump(all_labels,open(all_labels_path,'w'))
    diversity=len(set(all_labels)-set(all_goldens))
    all_tags_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/iter_labels/all_tags_5.json'
    # laod all tags
    all_tags=json.load(open(all_tags_path,'r'))
    diversity_all_tags=len(set(all_labels)-set(all_tags))
    print('diversity',diversity)
    print('diversity_all_tags',diversity_all_tags)
    centers_info=redo_15_10_clustring(all_labels,iter)
    silver_update_noveltyS(centers_info,silver_iter_train.format(iter),silver_iter_test.format(iter))
    print('redo_15_10_clustring done')
    # save all labels
    
# uodate silvers in iteratios (auto)    
def debug_silver_update(write_file_name='demo'):
    iter=0
    train_datas=json.load(open('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/test_iter1_0503_old_diva_dumping05.json'))
    test_datas=json.load(open('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/train_iter1_0503_old_diva_dumping05.json'))
    # train_datas=test_datas[0:100]
    # test_datas=test_datas[101:200]
    update_silver_iter(0,train_datas,test_datas,'golden','0508_baseline_diva_test_finetune_golden',write_in_xlsx=True,joint_score_thres=0.015,uncertainty_limit=True,update_thres=0.05,write_file_name=write_file_name)



def update_silver_simple(iter,train_datas,test_datas,ebc_thres):
    # if stop_flag=0 then stop iteration , otherwise continue
    # pro_new_label_songs=0
    stop_flag=0
    for data in train_datas:
        score_dict=data['final_score_dict_sort']
        sort_score_dict=sorted(score_dict.items(),key=lambda x:x[1][-1],reverse=True)
        silver_labels=[]
        for i in range(len(sort_score_dict)):
            if sort_score_dict[i][1][-1]>=ebc_thres:
                silver_labels.append(sort_score_dict[i][0])
            else:
                break
        if 'song_silver_labels' not in data or iter==0:
            data['song_silver_labels']=[]
        data['song_silver_labels'].append(silver_labels)
        if 'song_all_silver_labels' not in data or iter==0:
            data['song_all_silver_labels']=[]
        if len(set(silver_labels) - set(list(data['song_all_silver_labels'])+list(data['song_labels'])) )>=1:
            stop_flag+=1
        data['song_all_silver_labels'].extend(silver_labels)
        data['song_all_silver_labels']=list(set(data['song_all_silver_labels']))
    
    for data in test_datas:
        score_dict=data['final_score_dict_sort']
        sort_score_dict=sorted(score_dict.items(),key=lambda x:x[1][-1],reverse=True)
        silver_labels=[]
        for i in range(len(sort_score_dict)):
            if sort_score_dict[i][1][-1]>=ebc_thres:
                silver_labels.append(sort_score_dict[i][0])
            else:
                break
        if 'song_silver_labels' not in data or iter==0:
            data['song_silver_labels']=[]
        data['song_silver_labels'].append(silver_labels)
        if 'song_all_silver_labels' not in data or iter==0:
            data['song_all_silver_labels']=[]
        data['song_all_silver_labels'].extend(silver_labels)
        data['song_all_silver_labels']=list(set(data['song_all_silver_labels']))
    
    save_path_test='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/self-training/test_iter{}_golden'.format(iter)
    with open(save_path_test,'w') as f:
        json.dump(test_datas,f)
    
    print('update silver iter{} test done'.format(iter))
    
    """
    stop flag: 0 stop / others continue
    """
    print('stop_flag(pro new labels song)',stop_flag)
    return stop_flag,train_datas,test_datas
        

def update_silver_iter(iter,train_datas,test_datas,label_class,file_version,uncertainty_limit=False,with_tf_idf=True,with_novelty=True,with_valid=True,clusters_num=15,clustering_times=10,PE_thres=0.5,ebc_thres=0.95,tf_idf_type='old',all_goldens=golden_labels,object_test_f_path=object_test_f_path,object_train_f_path=object_train_f_path,write_in_xlsx=False,debug=False,joint_score_thres=0.0025,all_availeble_labels=all_available_labels,update_thres=0.05,dumping_weight=0.8,with_valid_semantic=True,with_valid_disc=True,only_joint=False,write_file_name='demo'):
    """
    input:
        iter: int 0,1,2,3,4...
        train_datas: list of dict
        test_datas:list of dict
        label class: expert or golden (baseline expert)
        file_version: save defferent version of silver labels and match score (for evaluate)
        uncertainty_limit: bool, if True, only update silver labels with uncertainty rank top 50%
        tf_idf_type:(str) old or new
        with_tf_idf: bool, if True, use tf_idf to update silver labels (for ablition study)
        with_novelty: bool, if True, use novelty to update silver labels (for ablition study)
        with_valid: bool, if True, use valid to update silver labels (for ablition study)
        clusters_num: int, number of clusters
        clustering_times: int, number of clustering times
        PE_thres: float, threshold of PE score
        ebc_thres: float, threshold of ebc score add to silver labels
        
        debug: bool, if True, only run 1000 datas and only for one epoch, and the thresholde set fto 0.5
        joint_score_thres: float, threshold of joint score
        write_file_name: str , the name of xlsx file(if write_in_xlsx=True)
        
    """
    # stop flag wether stop iteration -0 stop 1 continue
    print(label_class)
    stop_flag=0
    if debug:
        train_datas=train_datas[0:1000]
        test_datas=test_datas[0:100]
        ebc_thres=0.5
    # more_silver=0
    if label_class=='golden':
        silver_iter_test='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/test_iter{}_{}.json'.format(iter+1,file_version)
        silver_iter_train='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/train_iter{}_{}.json'.format(iter+1,file_version)
        all_labels_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/iter_labels/all_labels_iter{}_{}.json'.format(iter+1,file_version)
        ebc_dict_save_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/data_for_update_silver/ebc_dict_{}_{}.json'.format(iter+1,file_version)
        valid_dict_save_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/data_for_update_silver/valid_dict_{}_{}.json'.format(iter+1,file_version)
    elif label_class=='expert':
        silver_iter_test='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/test_iter{}_{}_expert.json'.format(iter+1,file_version)
        silver_iter_train='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/train_iter{}_{}_expert.json'.format(iter+1,file_version)
        all_labels_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/iter_labels/all_labels_iter{}_5expert.json'.format(iter+1,file_version)
    
   
    # print('diverse score analys over')
    
    if iter==0:
        ebc_s_dict=json.load(open('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/data_for_update_silver/ebc_dict_1_0508_baseline_diva_test_finetune_golden.json','r'))
    elif os.path.exists(ebc_dict_save_path):
        ebc_s_dict=json.load(open(ebc_dict_save_path,'r'))
        print('ebc dict is already exist')
    else:
        ebc_s_dict=get_ebc_score_distribution(train_datas+test_datas)
        with open(ebc_dict_save_path,'w') as f:
            json.dump(ebc_s_dict,f)
    print('ebc score distribution analys over')
    #todo debug time

    if iter==0:
        valid_score_dict=json.load(open('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/data_for_update_silver/valid_dict_1_0508_baseline_diva_test_finetune_golden.json','r'))
    elif os.path.exists(valid_dict_save_path):
        valid_score_dict=json.load(open(valid_dict_save_path,'r'))
        print('valid dict is already exist')
    else:
        valid_score_dict=get_valid_score_distribution(train_datas+test_datas,ebc_s_dict,label_freq)
        with open(valid_dict_save_path,'w') as f:
            json.dump(valid_score_dict,f)
    print('valid score distribution analys over')
    
    all_test_labels,all_test_pred_labels,test_datas=update_silver_iter1(test_datas,silver_iter_test.format(iter),iter,ebc_s_dict,valid_score_dict,save=False,label_class=label_class,uncertainty_limit=uncertainty_limit,tf_idf_type=tf_idf_type,with_novelty=with_novelty,with_tf_idf=with_tf_idf,with_valid=with_valid,PE_thres=PE_thres,ebc_thres=ebc_thres,write_in_xlsx=write_in_xlsx,joint_score_thres=joint_score_thres,update_thres=update_thres,dumping_weight=dumping_weight,with_valid_semantic=with_valid_semantic,with_valid_disc=with_valid_disc,only_joint=only_joint,write_file_name=write_file_name)
    print('update silver iter{} test done(no save)'.format(iter))
    all_train_labels,all_train_pred_labels,train_datas=update_silver_iter1(train_datas,silver_iter_train.format(iter),iter,ebc_s_dict,valid_score_dict,save=False,label_class=label_class,uncertainty_limit=uncertainty_limit,tf_idf_type=tf_idf_type,with_novelty=with_novelty,with_tf_idf=with_tf_idf,with_valid=with_valid,PE_thres=PE_thres,ebc_thres=ebc_thres,write_in_xlsx=False,joint_score_thres=joint_score_thres,update_thres=update_thres,dumping_weight=dumping_weight,with_valid_semantic=with_valid_semantic,with_valid_disc=with_valid_disc,only_joint=only_joint)
    print('update silver iter{} train done(no save)'.format(iter))

    
    # 只关注train是否出新东西 预测的095以上的标签
    all_labels_new=list(set(all_train_pred_labels))
    # silver + golden -下一次进入训练的label
    all_labels=list(set(all_train_labels+all_test_labels))
    # save all labels
    # 只对新标签进行聚类
    # json.dump(all_labels,open(all_labels_path,'w'))
    
    diversity=len(set(all_labels_new)-set(all_goldens))
    # tags 训练时采用的标签
    all_tags_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/iter_labels/all_tags.json'
    last_all_tags_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/iter_labels/all_tags_iter{}_{}_{}.json'.format(iter-1,file_version,label_class)
    save_all_tags_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/iter_labels/all_tags_iter{}_{}_{}.json'.format(iter,file_version,label_class)
    
    # laod all tags
    all_tags=json.load(open(all_tags_path,'r'))
    if iter==0:
        last_all_tags=list(set(all_available_labels) & set(all_tags))
    else:
        # load last iter all tags
        last_all_tags=json.load(open(last_all_tags_path,'r'))
    
    last_all_pred_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/iter_labels/all_tags_pred_iter{}_{}_{}.json'.format(iter-1,file_version,label_class)
    save_all_pred_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/iter_labels/all_tags_pred_iter{}_{}_{}.json'.format(iter,file_version,label_class)
    if iter==0:
        last_all_pred=[]
    else:
        #laod last iter all pred
        last_all_pred=json.load(open(last_all_pred_path,'r'))
        
    diversity_all_tags=len(set(all_labels_new)-set(all_tags))
    
    diversity_all_tags_2=len(set(all_labels)-set(all_tags))
    # save all pred 保存所有模型预测的标签
    json.dump(all_labels_new,open(save_all_pred_path,'w'))
    diversity_increment=len(set(all_labels_new)-set(last_all_pred))
    if diversity_increment != 0:
        stop_flag=1
        
    print('diversity(pred)',diversity)
    print('****diversity_all_tags(pred)*****',diversity_all_tags)
    print('****diversity_all_tags_2(update_silver)*****',diversity_all_tags_2)
    print('-iter-{}-diversity_increment:{}'.format(iter,diversity_increment))
    
    # save all labels to save_tag_path
    json.dump(all_labels,open(save_all_tags_path,'w'))
    print('save all labels to save_tag_path done')
    
    # add new centers info (当前扩增的用于聚类的标签数，重新进行聚类)
    labels_to_cluster=list(set(all_labels)-set(last_all_tags))
    # add_centers_num=int(len(labels_to_cluster)*(15/1000))
    if len(labels_to_cluster)<1000:
        add_centers_num=int(len(labels_to_cluster)*(15/1000))
    else:
        add_centers_num=15
    print('add_centers_num',add_centers_num)
    
    if iter==0:
        last_centers_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/iter_labels/centers_iter-1_basedexpert.npy'
    else:
        last_centers_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/iter_labels/centers_iter{}_{}_{}.npy'.format(iter-1,file_version,label_class)
    save_centers_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/iter_labels/centers_iter{}_{}_{}.npy'.format(iter,file_version,label_class)
    # load last centers info
    last_centers_info=np.load(last_centers_path,allow_pickle=True)
    # 查看last_centers_info的shape
    # print('last_centers_info.shape',last_centers_info.shape)
    if add_centers_num > 0:
        centers_info=redo_15_10_clustring(labels_to_cluster,iter,label_class=label_class,file_version=file_version,clusters_num=add_centers_num,clustering_times=clustering_times)
        # centers_info+=last_centers_info
        new_centers_info=[]
        for cluster_time in range(clustering_times):
            centers_info_tmp=np.vstack((last_centers_info[cluster_time],centers_info[cluster_time]))
            # print('centers_info_tmp.shape',centers_info_tmp.shape)
            new_centers_info.append(centers_info_tmp)
    # save centers infod
    else:
        new_centers_info=last_centers_info
    new_centers_info=np.array(new_centers_info)
    np.save(save_centers_path,new_centers_info)
    print('save centers info done ')
    
    train_datas,test_datas=silver_update_noveltyS(new_centers_info,train_datas,test_datas)
    print('redo_15_10_clustring done')
    # save test/train data list
    
    if debug==False:
        json.dump(train_datas,open(silver_iter_train,'w'))
        json.dump(test_datas,open(silver_iter_test,'w'))
        
    print('update silver iter{} done'.format(iter))
    print('-----------------------------------------------------------')
    return train_datas,test_datas,diversity_all_tags,stop_flag
    # save all labels

def init_novelty_score(centers_num=15,clustering_time=10,test_data_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/test_iter_based_expert.json',train_data_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/train_iter_based_expert.json'):
    # object_test_f_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/test_iter_based.json'
    # object_train_f_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/train_iter_based.json'
    
    object_test_f_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/test_iter_based_golden_{}_{}.json'.format(centers_num,clustering_time)
    object_train_f_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/train_iter_based_golden_{}_{}.json'.format(centers_num,clustering_time)
    # test_data_lst=json.load(open(test_data_path,'r'))
    # train_data_lst=json.load(open(train_data_path,'r'))
    test_data_lst=json.load(open('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/test_iter1_0503_old_diva_dumping05.json'))
    train_data_lst=json.load(open('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/train_iter1_0503_old_diva_dumping05.json'))
    print('data loaded')
    train_labels_dict=json.load(open('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/iter_labels/all_trained_labels(golden-expert).json','r'))
    # all_experts=train_labels_dict['expert']
    all_goldens=train_labels_dict['golden']
    
    # centers_info=np.load('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/iter_labels/centers_iter-1_basedexpert.npy',allow_pickle=True)
    centers_info=redo_15_10_clustring(all_goldens,iter=-1,label_class='expert',file_version='based',init=True,clusters_num=centers_num,clustering_times=clustering_time)
    centers_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/iter_labels/centers_iter{}_{}_{}.npy'.format(-1,file_version,label_class)
    train_datas,test_datas=silver_update_noveltyS(centers_info,train_data_lst,test_data_lst)
    json.dump(train_datas,open(object_train_f_path,'w'))
    json.dump(test_datas,open(object_test_f_path,'w'))
    print('expert novelty score init over')
    
    # centers_info=redo_15_10_clustring(all_goldens,iter=-1,label_class='golden',file_version='based',init=True,clusters_num=15,clustering_times=10)
    # train_datas,test_datas=silver_update_noveltyS(centers_info,train_data_lst,test_data_lst)
    # json.dump(train_datas,open(object_train_f_path,'w'))
    # json.dump(test_datas,open(object_test_f_path,'w'))
    # print('golden novelty score init over')
    
    
    
    
    
    
def redo_15_10_clustring(all_labels_lst,iter,label_class,file_version,clusters_num,clustering_times,init=False):
    """
    input: all_labels_lst ,num of clusters , clustering times
    output: all_labels_dict={labels:embedding}
    """
    if label_class=='golden':
        centers_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/iter_labels/centers_iter{}_{}.npy'.format(iter,file_version)
    else:
        centers_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/iter_labels/centers_iter{}_{}expert.npy'.format(iter,file_version)
    song_candidate_words_dict = np.load('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/tme_big_data/candidate_words_context_embedding_dict_1536_300_pca512.npy', allow_pickle=True).item()
    all_labels_dict={}
    for labels in all_labels_lst:
        if labels in song_candidate_words_dict:
            all_labels_dict[labels]=song_candidate_words_dict[labels]
    # kmeans
    centers_info=M_N_clustring(centers_path,list(all_labels_dict.values()),iter,clusters_num,clustering_times,init=init)
    print('centers vec updated...')
    return centers_info

def cal_f1(data_lst=None,label_freq=label_freq):
    """
    input : test data_lst
    output : f1
    """
    # song_id: annotated_labels
    p_lst=[]
    r_lst=[]
    f1_lst=[]
    # 标注数据量分�?
    # 用于生成
    annotated_num_lst=[]
    candiate_num_lst=[]
    golden_num_lst=[]
    # 评估缺失数量
    pram_lst=[]
    song_info_dict=json.load(open('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Evaluate/yao_check_datas/test_song_infos.json','r'))
    # not match valid score
    for data_item in data_lst:
        # todo all silver
        # all_silver=data_item['song_all_silver_labels']
        if data_item['song_name'] not in song_info_dict:
            continue
        annotated_labels=song_info_dict[data_item['song_name']]['annotated_labels']
        annotated_num_lst.append(len(annotated_labels))
        golden_labels=data_item['song_pseudo_golden_labels']
        all_labels=list(set(annotated_labels+golden_labels))
        golden_num_lst.append(len(golden_labels))
        pram_lst.append(len(annotated_labels)/len(all_labels))
        # p=len(set(all_silver)&set(annotated_labels))/len(set(all_silver))
        # r=len(set(all_silver)&set(annotated_labels))/len(set(annotated_labels))
        # todo
        score_dict=data_item['final_score_dict_sort']
        # golden_labels=data_item['song_pseudo_golden_labels']
        # sort by site -1
        score_dict_sort=sorted(score_dict.items(),key=lambda x:x[1][-1],reverse=True)
        # get candidate num
        candidate_num=len(score_dict_sort)
        candiate_num_lst.append(candidate_num)
        # get rank by site -1
        rank_dict={}
        max_val=1
        ebc_more_than_095=0
        # ebc 大于0.95 的silver
        ebc_lst=[]
        # for i in range(len(score_dict_sort)):
        #     if score_dict_sort[i][1][-1] < max_val:
        #         max_val=score_dict_sort[i][1][-1]
        #     if score_dict_sort[i][1][-1] >= 0.95:
        #         ebc_more_than_095+=1
        #         valid_score=label_freq[score_dict_sort[i][0]]['var']
        #         if valid_score >= 0.14:
        #             ebc_lst.append(score_dict_sort[i][0])
        #     rank_dict[score_dict_sort[i][0]]=i+1
        
    candiate_num_lst=np.array(candiate_num_lst)
    annotated_num_lst=np.array(annotated_num_lst)
    golden_num_lst=np.array(golden_num_lst)
    pram_lst=np.array(pram_lst)
    print('annotated_num_mean',annotated_num_lst.mean())
    print('candidate_num_mean',candiate_num_lst.mean())
    print('golden_num_mean',golden_num_lst.mean())
    print('pram_lst_mean',pram_lst.mean())
        #v silver num = mean(candiate num 8 0.05,ebc_more_than_095)
        # silver_num=int((candidate_num*0.05+ebc_more_than_095)/2)

    #     # label: [match ,novelty with match, tf-idf++ with match, valid]
    #     silver_dict={}
    #     for key in score_dict:
    #         match_score=score_dict[key][-1]
    #         novelty_match=math.pow(score_dict[key][12],2)+math.pow(score_dict[key][-1],2)
    #         freq_match=score_dict[key][4]/math.sqrt(rank_dict[key])
    #         valid_score=label_freq[key]['var']
    #         silver_dict[key]=[match_score,novelty_match,freq_match,valid_score]
    #     # sort by site 1 
    #     silver_dict_sort_site1=sorted(silver_dict.items(),key=lambda x:x[1][1],reverse=True)
    #     # sort by site 2
    #     silver_dict_sort_site2=sorted(silver_dict.items(),key=lambda x:x[1][2],reverse=True)

    #     # ebc_s 0.05~0.0001 valid_score >= 0.001
    #     novelty_match_lst=[]
    #     for i in range(len(silver_dict_sort_site1)):
    #         if silver_dict_sort_site1[i][1][1] >= 1:
    #             if silver_dict_sort_site1[i][1][0] <0.05 and silver_dict_sort_site1[i][1][0] > 0.0001 and silver_dict_sort_site1[i][1][3] >= 0.14 and is_number(silver_dict_sort_site1[i][0])==False and len(silver_dict_sort_site1[i][0])>1:
    #                 novelty_match_lst.append(silver_dict_sort_site1[i][0])
    #         else:
    #             break
        
    #     # ebc_s 0.05~0.0001 valid_score >= 0.001 top silver_num
    #     freq_match_lst=[]
    #     for i in range(silver_num):
    #         if silver_dict_sort_site2[i][1][0] <0.05 and silver_dict_sort_site2[i][1][0] > 0.0001 and silver_dict_sort_site2[i][1][3] >= 0.14 and is_number(silver_dict_sort_site2[i][0])==False and len(silver_dict_sort_site2[i][0])>1:
    #             freq_match_lst.append(silver_dict_sort_site2[i][0])
    # return 
        
# 打印所有silver labels
def print_silver_labels_workbook(data_lst):
    workbook=xlsxwriter.Workbook('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Evaluate/yao_check_datas/iter_check'+'0325修正标注.xlsx')
    worksheet = workbook.add_worksheet()
    blue_bold = workbook.add_format({'bold': True, 'font_color': '#9d2933'})
    red=workbook.add_format({'font_color': '#ff00ff'})
    green=workbook.add_format({'font_color': '#6b8e23'})
    yellow=workbook.add_format({'font_color': '#cfb53b'})
    bold=workbook.add_format({'bold': True})
    black=workbook.add_format({'font_color': 'black'})
    grey=workbook.add_format({'font_color': '#808080'})
    # add a format for Italic
    italic=workbook.add_format({'italic': True,'bg_color': '#fcefe6'})
    song_info_dict=json.load(open('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Evaluate/yao_check_datas/test_song_infos.json','r'))
    
    line=1
    for data in data_lst:
        song_name=data['song_name']
        golden_labels=data['song_pseudo_golden_labels']
        silver_labels=data['song_all_silver_labels']
        if song_name not in song_info_dict:
            continue
        annotated_labels=song_info_dict[song_name]['annotated_labels']
        # comments_lst=[usr1_comment,usr2_comment,usr3_comment...]
        annotated_candidate=list(set(annotated_labels+silver_labels+golden_labels))

        # todo write basic info
        worksheet.write(line,1,song_name)
        line+=1
        worksheet.write(line,1,'golden_labels')
        worksheet.write(line,2,' '.join(golden_labels))
        line+=1
        worksheet.write(line,1,'silver_labels')
        worksheet.write(line,2,' '.join(silver_labels))
        line+=1
        comments=data['song_comments_detail_final']
        comments_lst=[]
        for usr_id in comments:
            usr_comment=comments[usr_id]
            usr_comment_str=''
            for comment_s_id in usr_comment:
                if comment_s_id !='likecnt_weight':
                    usr_comment_str+=usr_comment[comment_s_id]['song_view_raw_replace_mask']
            comments_lst.append(usr_comment_str)
        for label in annotated_candidate:
            recent_color=black
            # if label in annotation_labels_test and label in pred_label_lst:
            #     worksheet.write(line,1,label,red)
            #     recent_color=red
            # elif label in annotation_labels_test and label not in pred_label_lst:
            #     worksheet.write(line,1,label,green)
            #     recent_color=green
            # elif label not in annotation_labels_test and label in pred_label_lst:
            #     worksheet.write(line,1,label,yellow)
            #     recent_color=yellow
            if label in silver_labels+golden_labels:
                # worksheet.write(line,1,label,bold)
                recent_color=bold
            # else
                # worksheet.write(line,1,label)
            # 判断label是否在哪条评论中
            star_line=line
            count=0
            cell_count=0
            for comment in comments_lst:
                if label in comment:
                    # write comment with label in red
                    cell_count+=1
                    starts=[each.start() for each in re.finditer(label,comment)]
                    args=[line,3]
                    site_i=0
                    while site_i <len(comment):
                        if site_i in starts:
                            args.append(blue_bold)
                            args.append(''.join(comment[site_i:site_i+len(label)]))
                            site_i+=len(label)
                        else:
                            args.append(comment[site_i])
                            site_i+=1
                    worksheet.write_rich_string(*args)
                    line+=1
            # merge cell
            if cell_count>1:
                worksheet.merge_range(star_line,2,line-1,2,cell_count)
                worksheet.merge_range(star_line,0,line-1,0,label,recent_color)
                worksheet.merge_range(star_line,1,line-1,1,'')
            else:
                worksheet.write(star_line,2,cell_count)
                worksheet.write(star_line,0,label,recent_color)
        line+=1
    workbook.close()
    # workbook.save('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Evaluate/yao_check_datas/iter_test_pred_silver_labels.xlsx')


#todo 暂时不用
def update_silver_iter2(f_path,save_f_path,all_goldens=golden_labels):
    workbook=openpyxl.Workbook()
    woorksheet=workbook.active
    data_lst=json.load(open(f_path,'r'))
    all_labels=[]
    line=1
    time=1
    for data_item in data_lst:
        #get silver labels �? ebc_based >=0.95 + divers_iter0扩充 �? duivers_iter2扩充 （ebc_iter1>0.9 and ebc_iter1-ebc_based=0)
        time+=1
        if time>=100:
            break
        ebc_based=[]
        song_name=data_item['song_name']
        golden_labels=data_item['song_pseudo_golden_labels']
        silver_iter0=data_item['song_siver_labels']
        woorksheet.cell(line,1,song_name)
        line+=1
        woorksheet.cell(line,1,'golden')
        woorksheet.cell(line,2,' '.join(golden_labels))
        line+=1
        woorksheet.cell(line,1,'silver_iter0')
        woorksheet.cell(line,2,' '.join(silver_iter0))
        line+=1
        diverse_iter2=[]
        score_dict=data_item['final_score_dict_sort']
        # sort by site -1
        score_dict_sort_iter1=sorted(score_dict.items(),key=lambda x:x[1][-1],reverse=True)
        # sort by site -2
        score_dict_sort_iter0=sorted(score_dict.items(),key=lambda x:x[1][-2],reverse=True)
        for item in score_dict_sort_iter0:
            if item[1][-2] >= 0.95:
                ebc_based.append(item[0])
        
        for item in score_dict_sort_iter1:
            if item[1][-1] >= 0.9 and item[1][-1]-item[1][-2] >=0.9:
                diverse_iter2.append(item[0])
        
        silver_labels=list(set(ebc_based)|(set(silver_iter0)&set(diverse_iter2)))
        woorksheet.cell(line,1,'silver_iter2')
        woorksheet.cell(line,2,' '.join(silver_labels))
        line+=2
        data_item['song_siver_labels']=silver_labels
        all_labels+=(silver_labels+golden_labels)
    # save to file
    # json.dump(data_lst,open(save_f_path,'w'))
    # diversity
    dicersity=len(set(all_labels)-set(all_goldens))
    print('diversity',dicersity)
    workbook.save('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Evaluate/yao_check_datas/iter_check/iter_silvers.xlsx')

iter1_test_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/eval_data/test_ebc_val_iter1_0320_nodropout.json'
iter1_train_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/eval_data/train_ebc_val_iter1_0320_nodropout.json'
iter2_test_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/test_t0_for_supervise_DIVA_0322_silver2.json'
iter2_train_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/train_t0_for_supervise_DIVA_0322_silver2.json'

# update_silver_iter2(iter1_train_path,iter2_train_path)


    
if '__main__'==__name__:
    write_file_name='0517_base'
    debug_silver_update(write_file_name)
    
    # debug_silver_update()
    init_novelty_score(clustering_time=1)
    init_novelty_score(clustering_time=5)
    init_novelty_score(clustering_time=15)
    
    init_novelty_score(centers_num=5)
    init_novelty_score(centers_num=10)
    init_novelty_score(centers_num=20)
    # print('here updating silver labels')
    # iter=1
    # update_silver(iter)
    # silver_iter_test='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/test_iter{}_4.json'
    # silver_iter_train='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/train_iter{}_4.json'
    # all_labels=json.load(open('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/iter_labels/all_labels_iter{}_4.json'.format(iter),'r'))
    # all_tags_path='/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/iter_labels/all_tags.json'
    # # laod all tags
    # all_tags=json.load(open(all_tags_path,'r'))
    # diversity=len(set(all_labels)-set(all_tags))
    # print('diversity',diversity)
    # centers_info=redo_15_10_clustring(all_labels,iter)
    # silver_update_noveltyS(centers_info,silver_iter_train.format(iter),silver_iter_test.format(iter))

    #todo annotation analysis
    # print(cal_f1(json.load(open('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/test_iter{}.json'.format(3),'r'))))
    # print_silver_labels_workbook(json.load(open('/home/yaoyao123/yaoyao123_data/keyphrasegenration/keyphrase-classification-New-Architecture-New-Data-Split/Preprocess/yao_preprocess_data/diva_data/test_iter{}.json'.format(3),'r')))

















# update_silver_iter1(train_f_path,silver_iter1_train)
# from transformers import pipeline
        
        
        
